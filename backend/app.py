import csv
import io
import json
import os
import re
import base64
import hashlib
import hmac
import sys
import shutil
from datetime import datetime
from functools import wraps
from urllib import error as urllib_error
from urllib import request as urllib_request


# Desktop/AppData-safe runtime paths. Must be set before utils.database imports.
def _configure_desktop_runtime_paths():
    try:
        appdata = os.getenv("APPDATA") or os.path.expanduser("~")
        base = os.path.join(appdata, "HospAI")
        logs_dir = os.path.join(base, "logs")
        os.makedirs(logs_dir, exist_ok=True)
        os.environ.setdefault("SESSION_COOKIE_SAMESITE", "Lax")
        os.environ.setdefault("SESSION_COOKIE_SECURE", "false")
    except Exception:
        pass

_configure_desktop_runtime_paths()

from dotenv import load_dotenv
from flask import Flask, g, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from fpdf import FPDF
from utils.auth import (
    ADMIN_ROUTE_AUTH_COOKIE_NAME,
    ADMIN_ROUTE_AUTH_TTL_SECONDS,
    ASSIGNABLE_MODULES,
    SESSION_COOKIE_NAME,
    SESSION_TTL_SECONDS,
    authenticate,
    check_username_exists,
    create_admin_route_auth_token,
    create_default_users,
    create_session,
    delete_session,
    get_session_user,
    is_admin_route_auth_configured,
    reset_hospital_admin_password,
    resolve_user_permissions,
    signup_employee,
    signup_hospital_admin,
    verify_admin_route_auth_token,
    verify_admin_route_password,
)
from utils.export import draw_pdf_brand_header, generate_pdf, generate_word
from utils.ocr import LANGUAGE_NAMES, extract_text_from_image
from utils.storage import ObjectStorage
from werkzeug.exceptions import BadRequest
from werkzeug.utils import secure_filename

from utils.database import (
    activate_employee,
    add_admission,
    add_audit_log,
    add_document,
    add_medication_schedule,
    add_observation_note,
    add_patient,
    add_patient_movement,
    assign_bed,
    check_duplicate_patient,
    create_account_ledger_entry,
    create_attendance,
    create_appointment,
    create_certificate,
    create_department,
    create_diagnostic_record,
    create_doctor_schedule,
    create_doctor_payout,
    create_encounter,
    create_hospital,
    create_shared_export,
    get_shared_export,
    create_insurance_verification,
    create_insurance_claim,
    create_invoice,
    create_lab_vendor,
    create_leave_request,
    create_ot_surgery,
    create_ot_theatre,
    create_patient_consent,
    create_payroll_record,
    create_vendor_payment,
    current_ist_datetime,
    delete_account_ledger_entry,
    deactivate_employee,
    delete_attendance_record,
    delete_certificate,
    delete_department,
    delete_diagnostic_record,
    delete_doctor_schedule,
    delete_doctor_payout,
    delete_document,
    delete_employee,
    delete_insurance_claim,
    delete_invoice,
    delete_lab_vendor,
    delete_leave_request,
    delete_ot_surgery,
    delete_ot_theatre,
    delete_patient,
    delete_payroll_record,
    delete_vendor_payment,
    generate_patient_id,
    get_admissions,
    get_accounts_summary,
    get_all_employees,
    get_all_patients,
    get_audit_logs,
    get_dashboard_analytics,
    get_diagnostic_summary,
    get_document,
    get_documents,
    get_employee,
    get_employee_stats,
    get_hospital_by_code,
    get_hospital_dashboard_summary,
    get_connection,
    get_invoice_by_id,
    get_op_summary,
    get_ot_summary,
    get_patient,
    get_patient_by_phone,
    get_patient_stats,
    get_reports_overview,
    get_revenue_summary,
    get_report_patient_history,
    init_database,
    list_account_ledger_entries,
    list_attendance,
    list_appointments,
    list_doctors_history,
    list_bed_allocations,
    list_certificates,
    list_departments,
    list_diagnostics,
    list_doctor_schedules,
    list_doctor_payouts,
    list_insurance_claims,
    list_insurance_verifications,
    list_encounters,
    list_hospitals,
    list_invoices,
    list_lab_vendors,
    list_leave_requests,
    list_medication_schedules,
    list_observation_notes,
    list_ot_surgeries,
    list_ot_theatres,
    list_patient_consents,
    list_patient_movements,
    list_payroll,
    list_vendor_payments,
    record_invoice_payment,
    resolve_hospital_id,
    search_employees,
    search_patients,
    set_hospital_status,
    update_account_ledger_entry,
    update_attendance_record,
    update_appointment,
    update_department,
    update_diagnostic_record,
    update_doctor_schedule,
    update_doctor_payout,
    update_document_ocr,
    update_employee,
    update_insurance_verification,
    update_insurance_claim,
    update_invoice,
    update_lab_vendor,
    update_leave_status,
    update_ot_surgery,
    update_ot_theatre,
    update_patient_consent,
    update_patient,
    update_payroll_record,
    update_vendor_payment,
)

BASE_DIR = os.path.dirname(__file__)
PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, '..'))
load_dotenv(os.path.join(PROJECT_ROOT, '.env'), override=False)
load_dotenv(os.path.join(BASE_DIR, '.env'), override=False)
FRONTEND_DIST = os.path.abspath(os.path.join(PROJECT_ROOT, 'frontend', 'dist'))
STORAGE = ObjectStorage()

app = Flask(__name__)
default_allowed_origins = {
    "https://app.hospai.ai",
    "https://staging-app.hospai.ai",
    "http://localhost:5173",
    "http://127.0.0.1:5173",
}
env_allowed_origins = {
    origin.strip()
    for origin in os.getenv("CORS_ALLOWED_ORIGINS", "").split(",")
    if origin.strip()
}
is_development = os.getenv("FLASK_ENV", "").lower() == "development"
ALLOWED_ORIGINS = set(default_allowed_origins)
ALLOWED_ORIGINS.update(env_allowed_origins)


def is_allowed_origin(origin: str | None) -> bool:
    # Electron production uses file:// for the renderer. Chromium sends
    # Origin: null for file:// -> http://127.0.0.1 API requests.
    # Allowing this exact origin is required for the login response to be
    # readable by the desktop UI.
    if origin == "null":
        return True
    if not origin:
        return False
    if origin in ALLOWED_ORIGINS:
        return True
    # Allow all HospAI HTTPS subdomains (staging/prod/custom app surfaces).
    if re.match(r"^https://([a-z0-9-]+\.)*hospai\.ai$", origin):
        return True
    # Allow localhost and private LAN dev servers on any port (e.g. Vite 5173/5174/4173).
    return re.match(r"^https?://(localhost|127\.0\.0\.1|192\.168\.\d+\.\d+|10\.\d+\.\d+\.\d+|172\.(1[6-9]|2\d|3[0-1])\.\d+\.\d+)(:\d+)?$", origin) is not None


CORS(
    app,
    resources={
        r"/api/*": {
            "origins": [
                r"^https://([a-z0-9-]+\.)*hospai\.ai$",
                "null",
                r"^https?://(localhost|127\.0\.0\.1|192\.168\.\d+\.\d+|10\.\d+\.\d+\.\d+|172\.(1[6-9]|2\d|3[0-1])\.\d+\.\d+)(:\d+)?$",
                *sorted(ALLOWED_ORIGINS),
            ]
        }
    },
    supports_credentials=True,
    allow_headers=[
        "Content-Type",
        "Authorization",
        "X-Requested-With",
        "X-Hospital-Code",
        "X-Platform-Admin-Username",
        "X-Platform-Admin-Password",
    ],
    methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
)


@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin")
    if is_allowed_origin(origin):
        response.headers.setdefault("Access-Control-Allow-Origin", origin)
        response.headers.setdefault("Access-Control-Allow-Credentials", "true")
        response.headers.setdefault(
            "Access-Control-Allow-Headers",
            "Content-Type, Authorization, X-Requested-With, X-Hospital-Code, X-Platform-Admin-Username, X-Platform-Admin-Password",
        )
        response.headers.setdefault(
            "Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS"
        )
    return response


@app.route("/api/<path:_path>", methods=["OPTIONS"])
def preflight(_path):
    return ("", 204)


def _desktop_license_base_dir() -> str:
    appdata = os.getenv("APPDATA") or os.path.expanduser("~")
    return os.path.join(appdata, "HospAI")


def _desktop_license_log(message: str) -> None:
    try:
        logs_dir = os.path.join(_desktop_license_base_dir(), "logs")
        os.makedirs(logs_dir, exist_ok=True)
        with open(os.path.join(logs_dir, "license.log"), "a", encoding="utf-8") as handle:
            handle.write(f"{datetime.now().isoformat()} backend_guard {message}\n")
    except Exception:
        pass


def _desktop_license_guard_enabled() -> bool:
    value = (os.getenv("HOSPAI_ENFORCE_LICENSE") or "").strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    # Packaged backend executables must not serve APIs by themselves unless
    # Electron starts them with a verified license environment.
    return bool(getattr(sys, "frozen", False))


def _read_json_file(path: str) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
            return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def _desktop_license_error() -> str | None:
    if not _desktop_license_guard_enabled():
        return None

    base = _desktop_license_base_dir()
    block_path = os.path.join(base, "licenses", "block.json")
    block_payload = _read_json_file(block_path)
    if block_payload.get("blocked") is True:
        _desktop_license_log("manual_block_active")
        return "Access blocked. Please contact Kalpra Tech Solutions."

    registry_path = os.path.join(base, "licenses", "license_registry.json")
    registry = _read_json_file(registry_path)
    last_seen = str(registry.get("last_seen_at") or "").strip()
    if last_seen:
        try:
            last_seen_dt = datetime.fromisoformat(last_seen.replace("Z", "+00:00"))
            now_dt = datetime.now(last_seen_dt.tzinfo) if last_seen_dt.tzinfo else datetime.now()
            if now_dt.timestamp() + (5 * 60) < last_seen_dt.timestamp():
                _desktop_license_log(f"time_rollback now={now_dt.isoformat()} last_seen={last_seen}")
                return "System time changed. Access blocked. Please contact Kalpra Tech Solutions."
        except Exception:
            pass

    expiry_raw = (os.getenv("HOSPAI_LICENSE_EXPIRY_TS") or "").strip()
    if not expiry_raw:
        _desktop_license_log("missing_verified_expiry_env")
        return "Offline license validation missing. Please open HospAI using the desktop launcher."
    try:
        expiry_ms = int(float(expiry_raw))
    except ValueError:
        _desktop_license_log(f"invalid_verified_expiry_env={expiry_raw}")
        return "License validation missing. Please contact Kalpra Tech Solutions."
    now_ms = int(datetime.now().timestamp() * 1000)
    if now_ms > expiry_ms:
        display = os.getenv("HOSPAI_LICENSE_EXPIRY_DISPLAY") or str(expiry_ms)
        _desktop_license_log(f"expired now_ms={now_ms} expiry={display}")
        return "Time has expired. Please contact Kalpra Tech Solutions."
    return None


@app.before_request
def enforce_desktop_license_guard():
    # Health remains available so Electron can confirm that the backend process started.
    if request.method == "OPTIONS" or request.path == "/api/health":
        return None
    if not request.path.startswith("/api/"):
        return None
    error_message = _desktop_license_error()
    if error_message:
        return jsonify({"success": False, "error": error_message}), 403
    return None


init_database()
create_default_users()


def row_to_dict(row):
    if row is None:
        return None
    data = dict(row)
    module_access = data.get("module_access")
    if isinstance(module_access, str):
        try:
            parsed = json.loads(module_access)
            if isinstance(parsed, list):
                data["module_access"] = parsed
        except json.JSONDecodeError:
            data["module_access"] = []
    return data


def rows_to_dicts(rows):
    return [row_to_dict(row) for row in rows]


PLATFORM_ADMIN_USERNAME = (os.getenv("PLATFORM_ADMIN_USERNAME") or "").strip()
PLATFORM_ADMIN_PASSWORD = (os.getenv("PLATFORM_ADMIN_PASSWORD") or "").strip()
RAZORPAY_KEY_ID = (os.getenv("RAZORPAY_KEY_ID") or "").strip()
RAZORPAY_KEY_SECRET = (os.getenv("RAZORPAY_KEY_SECRET") or "").strip()
RAZORPAY_API_BASE = "https://api.razorpay.com/v1"


def validate_hospital_code(value: str) -> str:
    code = (value or "").strip().lower()
    if not code:
        raise BadRequest("hospital_code is required")
    if not re.fullmatch(r"[a-z0-9-]{3,64}", code):
        raise BadRequest("hospital_code must contain only lowercase letters, numbers, and '-'")
    return code


def request_hospital_id() -> int:
    requested_code = request.headers.get("X-Hospital-Code", "")
    try:
        code = validate_hospital_code(requested_code) if requested_code else None
    except BadRequest:
        code = None
    return resolve_hospital_id(code)


def current_hospital_id() -> int:
    user = getattr(g, "current_user", None) or {}
    return int(user.get("hospital_id") or request_hospital_id())


def require_platform_admin():
    username = (
        request.headers.get("X-Platform-Admin-Username")
        or request.args.get("platform_admin_username")
        or (request.get_json(silent=True) or {}).get("platform_admin_username")
        or ""
    ).strip()
    password = (
        request.headers.get("X-Platform-Admin-Password")
        or request.args.get("platform_admin_password")
        or (request.get_json(silent=True) or {}).get("platform_admin_password")
        or ""
    )

    _admin_user = (os.getenv("PLATFORM_ADMIN_USERNAME") or os.getenv("ONBOARDING_ADMIN_USERNAME") or "").strip()
    _admin_pass = (os.getenv("PLATFORM_ADMIN_PASSWORD") or os.getenv("ONBOARDING_ADMIN_PASSWORD") or "").strip()

    if not username and not password:
        if not _admin_user or not _admin_pass:
            return jsonify({"error": "Platform admin credentials are not configured"}), 503
        return jsonify({"error": "Platform admin authentication failed"}), 403

    if not _admin_user or not _admin_pass or username != _admin_user or password != _admin_pass:
        return jsonify({"error": "Platform admin authentication failed"}), 403
    return None


def _session_cookie_settings():
    secure_cookie = (
        request.is_secure or os.getenv("SESSION_COOKIE_SECURE", "").lower() == "true"
    )
    same_site = os.getenv("SESSION_COOKIE_SAMESITE", "Lax")
    if same_site not in ("Lax", "Strict", "None"):
        same_site = "Lax"
    return {
        "httponly": True,
        "secure": secure_cookie,
        "samesite": same_site,
        "path": "/",
        "max_age": SESSION_TTL_SECONDS,
    }


def _request_session_token():
    token = request.cookies.get(SESSION_COOKIE_NAME)
    if token:
        return token
    auth_header = request.headers.get("Authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return None


def require_session(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        token = _request_session_token()
        user = get_session_user(token)
        if not user:
            return jsonify({"error": "Authentication required"}), 401
        g.current_user = user
        return view(*args, **kwargs)

    return wrapper


def require_permissions(*required_permissions):
    def decorator(view):
        @wraps(view)
        @require_session
        def wrapper(*args, **kwargs):
            user_permissions = resolve_user_permissions(g.current_user)
            if not all(
                permission in user_permissions for permission in required_permissions
            ):
                return (
                    jsonify(
                        {
                            "error": "Forbidden",
                            "required_permissions": list(required_permissions),
                        }
                    ),
                    403,
                )
            return view(*args, **kwargs)

        return wrapper

    return decorator


def save_uploaded_file(uploaded_file, file_bytes, doc_type):
    original_name = secure_filename(uploaded_file.filename or "") or "document"
    return STORAGE.store(doc_type, original_name, file_bytes, uploaded_file.mimetype)


def log_audit_event(action, module_name, entity_key=None, payload=None):
    actor = (getattr(g, "current_user", None) or {}).get("username")
    serialized_payload = None
    if payload is not None:
        try:
            serialized_payload = json.dumps(payload, separators=(",", ":"), default=str)
        except Exception:
            serialized_payload = str(payload)
    add_audit_log(
        {
            "actor_username": actor,
            "action": action,
            "module_name": module_name,
            "entity_key": entity_key,
            "payload": serialized_payload,
            "ip_address": request.headers.get("X-Forwarded-For", request.remote_addr),
        }
    )


def validate_required_fields(payload, fields):
    missing = [field for field in fields if payload.get(field) in (None, "")]
    if missing:
        return jsonify({"error": "Missing required fields", "missing": missing}), 400
    return None


def is_razorpay_configured():
    return bool(RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET)


def require_razorpay_configured():
    if is_razorpay_configured():
        return None
    return jsonify({"error": "Razorpay is not configured. Set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET."}), 503


def to_amount_paise(value):
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    if amount <= 0:
        return None
    return int(round(amount * 100))


def _razorpay_auth_header():
    encoded = base64.b64encode(f"{RAZORPAY_KEY_ID}:{RAZORPAY_KEY_SECRET}".encode("utf-8")).decode("ascii")
    return f"Basic {encoded}"


def create_razorpay_order(amount_paise, currency="INR", receipt=None, notes=None):
    payload = {
        "amount": int(amount_paise),
        "currency": currency or "INR",
    }
    if receipt:
        payload["receipt"] = str(receipt)
    if isinstance(notes, dict) and notes:
        payload["notes"] = notes

    req = urllib_request.Request(
        f"{RAZORPAY_API_BASE}/orders",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": _razorpay_auth_header(),
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib_request.urlopen(req, timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib_error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            payload = {}
        message = (
            payload.get("error", {}).get("description")
            or payload.get("error", {}).get("code")
            or body
            or "Razorpay order creation failed."
        )
        raise RuntimeError(message) from exc
    except Exception as exc:
        raise RuntimeError("Unable to reach Razorpay.") from exc


def verify_razorpay_signature(order_id, payment_id, signature):
    signed_payload = f"{order_id}|{payment_id}".encode("utf-8")
    expected = hmac.new(
        RAZORPAY_KEY_SECRET.encode("utf-8"),
        signed_payload,
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected, str(signature or ""))


def normalize_payment_mode(value):
    allowed_modes = {"cash", "card", "upi", "bank"}
    mode = (value or "").strip().lower()
    return mode if mode in allowed_modes else "upi"


def normalize_billing_module(value):
    text = str(value or "").strip().lower()
    if "pharm" in text or "medicine" in text:
        return "PHARMACY"
    if "lab" in text or "diagnostic" in text or "test" in text:
        return "LAB"
    if text == "ip" or "ipd" in text or "admission" in text:
        return "IP"
    return "OP"


def normalize_department_name(value):
    normalized = " ".join(str(value or "").strip().split())
    if not normalized:
        raise BadRequest("department_name is required")
    if len(normalized) > 120:
        raise BadRequest("department_name must be at most 120 characters")
    return normalized


@app.get("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.get("/api/payments/razorpay/config")
def razorpay_config():
    return jsonify(
        {
            "configured": is_razorpay_configured(),
            "key_id": RAZORPAY_KEY_ID if is_razorpay_configured() else "",
        }
    )


@app.route("/api/auth/login", methods=["POST", "OPTIONS"])
def login():
    if request.method == "OPTIONS":
        return ("", 204)
    payload = request.get_json(force=True)
    hospital_id = request_hospital_id()
    user = authenticate(
        payload.get("username", ""),
        payload.get("password", ""),
        hospital_id=hospital_id,
    )
    if not user:
        return jsonify({"error": "Invalid credentials"}), 401
    if "error" in user:
        return jsonify(user), 403
    session_token, _expires_at = create_session(
        user_id=user["id"],
        hospital_id=user["hospital_id"],
        ip_address=request.headers.get("X-Forwarded-For", request.remote_addr),
        user_agent=request.headers.get("User-Agent"),
    )
    response = jsonify({"user": {k: v for k, v in user.items() if k != "id"}, "session_token": session_token})
    response.headers["Cache-Control"] = "no-store"
    response.set_cookie(
        SESSION_COOKIE_NAME, session_token, **_session_cookie_settings()
    )
    return response


@app.post("/api/auth/signup")
def signup():
    payload = request.get_json(force=True)
    result = signup_employee(payload, allow_admin_creation=False, hospital_id=request_hospital_id())
    status = 201 if result.get("success") else 400
    return jsonify(result), status


@app.get("/api/auth/check-username")
def check_username():
    username = request.args.get("username")
    if not username:
        return jsonify({"available": False, "error": "username required"}), 400
    available = not check_username_exists(username, hospital_id=request_hospital_id())
    return jsonify({"available": available})


@app.route("/api/auth/setup-admin", methods=["POST", "OPTIONS"])
def setup_hospital_admin():
    if request.method == "OPTIONS":
        return ("", 204)
    admin_gate = require_platform_admin()
    if admin_gate:
        return admin_gate
    hospital_id = request_hospital_id()
    payload = request.get_json(force=True)
    result = signup_hospital_admin(payload, hospital_id=hospital_id)
    if not result.get("success"):
        message = (result.get("message") or "").lower()
        status = 409 if "already configured" in message else 400
        return jsonify(result), status

    user = authenticate(
        payload.get("username", ""),
        payload.get("password", ""),
        hospital_id=hospital_id,
    )
    if not user:
        return jsonify(
            {"success": False, "message": "Admin created but automatic login failed."}
        ), 500

    session_token, _expires_at = create_session(
        user_id=user["id"],
        hospital_id=user["hospital_id"],
        ip_address=request.headers.get("X-Forwarded-For", request.remote_addr),
        user_agent=request.headers.get("User-Agent"),
    )
    response = jsonify(
        {
            "success": True,
            "message": result["message"],
            "user": {k: v for k, v in user.items() if k != "id"},
        }
    )
    response.headers["Cache-Control"] = "no-store"
    response.set_cookie(
        SESSION_COOKIE_NAME, session_token, **_session_cookie_settings()
    )
    return response, 201


@app.get("/api/platform/hospitals")
def platform_hospitals_list():
    admin_gate = require_platform_admin()
    if admin_gate:
        return admin_gate
    hospitals = list_hospitals()
    return jsonify({"hospitals": rows_to_dicts(hospitals)})


@app.post("/api/platform/hospitals")
def platform_hospitals_create():
    admin_gate = require_platform_admin()
    if admin_gate:
        return admin_gate
    payload = request.get_json(force=True)
    hospital_code = validate_hospital_code(payload.get("hospital_code", ""))
    hospital_name = payload.get("name")
    hospital_id, created = create_hospital(hospital_code, hospital_name)
    hospital = get_hospital_by_code(hospital_code)
    status = 201 if created else 200
    return jsonify(
        {
            "created": created,
            "hospital_id": hospital_id,
            "hospital": row_to_dict(hospital),
        }
    ), status


@app.post("/api/platform/hospitals/<hospital_code>/disable")
def platform_hospital_disable(hospital_code):
    admin_gate = require_platform_admin()
    if admin_gate:
        return admin_gate
    payload = request.get_json(silent=True) or {}
    code = validate_hospital_code(hospital_code)
    changed = set_hospital_status(code, "inactive", reason=payload.get("reason"))
    if not changed:
        return jsonify({"error": "Hospital not found"}), 404
    return jsonify(
        {"success": True, "hospital": row_to_dict(get_hospital_by_code(code))}
    )


@app.post("/api/platform/hospitals/<hospital_code>/enable")
def platform_hospital_enable(hospital_code):
    admin_gate = require_platform_admin()
    if admin_gate:
        return admin_gate
    code = validate_hospital_code(hospital_code)
    changed = set_hospital_status(code, "active")
    if not changed:
        return jsonify({"error": "Hospital not found"}), 404
    return jsonify(
        {"success": True, "hospital": row_to_dict(get_hospital_by_code(code))}
    )


@app.post("/api/platform/hospitals/<hospital_code>/admin/reset-password")
def platform_admin_reset_password(hospital_code):
    admin_gate = require_platform_admin()
    if admin_gate:
        return admin_gate

    code = validate_hospital_code(hospital_code)
    hospital = get_hospital_by_code(code)
    if not hospital:
        return jsonify({"error": "Hospital not found"}), 404

    payload = request.get_json(force=True)
    username = payload.get("username", "")
    new_password = payload.get("new_password", "")
    result = reset_hospital_admin_password(hospital["id"], username, new_password)
    status = 200 if result.get("success") else 400
    return jsonify(result), status


@app.get("/api/auth/session")
def auth_session():
    token = _request_session_token()
    user = get_session_user(token)
    if not user:
        response = jsonify({"user": None})
        response.headers["Cache-Control"] = "no-store"
        return response
    response = jsonify({"user": user})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.post("/api/auth/logout")
def auth_logout():
    token = _request_session_token()
    delete_session(token)
    response = jsonify({"success": True})
    response.headers["Cache-Control"] = "no-store"
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    response.delete_cookie(ADMIN_ROUTE_AUTH_COOKIE_NAME, path="/")
    return response


def require_admin_route_auth(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        token = request.cookies.get(ADMIN_ROUTE_AUTH_COOKIE_NAME)
        if not verify_admin_route_auth_token(token):
            return jsonify({"error": "Admin route authentication required"}), 403
        return view(*args, **kwargs)

    return wrapper


@app.get("/api/admin/auth/session")
def admin_route_auth_session():
    token = request.cookies.get(ADMIN_ROUTE_AUTH_COOKIE_NAME)
    authorized = verify_admin_route_auth_token(token)
    return jsonify(
        {"authorized": authorized, "configured": is_admin_route_auth_configured()}
    )


@app.post("/api/admin/auth/login")
def admin_route_auth_login():
    if not is_admin_route_auth_configured():
        return jsonify(
            {"error": "ADMIN_ROUTE_PASSWORD is not configured on the server."}
        ), 503

    payload = request.get_json(force=True)
    password = payload.get("password", "")
    if not verify_admin_route_password(password):
        return jsonify({"error": "Invalid admin route password."}), 403

    token = create_admin_route_auth_token()
    response = jsonify({"success": True, "authorized": True})
    response.set_cookie(
        ADMIN_ROUTE_AUTH_COOKIE_NAME,
        token,
        httponly=True,
        secure=request.is_secure
        or os.getenv("SESSION_COOKIE_SECURE", "").lower() == "true",
        samesite=os.getenv("SESSION_COOKIE_SAMESITE", "Lax"),
        path="/",
        max_age=ADMIN_ROUTE_AUTH_TTL_SECONDS,
    )
    return response


@app.post("/api/admin/auth/logout")
def admin_route_auth_logout():
    response = jsonify({"success": True})
    response.delete_cookie(ADMIN_ROUTE_AUTH_COOKIE_NAME, path="/")
    return response


@app.post("/api/admin/create-account")
@require_admin_route_auth
def admin_create_account():
    payload = request.get_json(force=True) or {}
    forced_payload = {
        **payload,
        "user_type": "admin",
        "module_access": [
            "dashboard",
            "patients",
            "billing",
            "lab",
            "hrms",
        ],
    }
    result = signup_employee(forced_payload, allow_admin_creation=True, hospital_id=current_hospital_id())
    status = 201 if result.get("success") else 400
    return jsonify(result), status


@app.get("/api/admin/users")
@require_admin_route_auth
def admin_users_list():
    return jsonify({"users": rows_to_dicts(get_all_employees())})


@app.post("/api/admin/users")
@require_admin_route_auth
def admin_users_create():
    payload = request.get_json(force=True) or {}
    requested_type = str(payload.get("user_type", "normal")).strip().lower()
    user_type = "admin" if requested_type == "admin" else "normal"
    module_access = (
        list(ASSIGNABLE_MODULES)
        if user_type == "admin"
        else payload.get("module_access", [])
    )

    created_payload = {
        **payload,
        "user_type": user_type,
        "module_access": module_access,
    }
    result = signup_employee(created_payload, allow_admin_creation=True, hospital_id=current_hospital_id())
    status = 201 if result.get("success") else 400
    return jsonify(result), status


@app.post("/api/admin/users/<employee_id>/promote")
@require_admin_route_auth
def admin_promote_user(employee_id):
    target = get_employee(employee_id=employee_id)
    if not target:
        return jsonify({"error": "Employee not found"}), 404

    promoted = update_employee(
        employee_id,
        {
            "user_type": "admin",
            "access_role": "owner",
            "module_access": list(ASSIGNABLE_MODULES),
        },
    )
    if not promoted:
        return jsonify({"error": "Employee not found"}), 404

    return jsonify({"success": True, "employee_id": employee_id, "user_type": "admin"})


@app.get("/api/languages")
def languages():
    return jsonify({"languages": LANGUAGE_NAMES})


@app.get("/api/reports/overview")
@require_permissions("reports.read")
def reports_overview():
    return jsonify(get_reports_overview())


@app.get("/api/reports/revenue-summary")
@require_permissions("reports.read")
def reports_revenue_summary():
    collection_date = request.args.get("date")
    return jsonify(get_revenue_summary(collection_date=collection_date))


@app.get("/api/reports/patient-history")
@require_permissions("reports.read")
def reports_patient_history():
    collection_date = request.args.get("date")
    filter_type = request.args.get("type")
    filter_value = request.args.get("value")
    return jsonify(get_report_patient_history(collection_date=collection_date, filter_type=filter_type, filter_value=filter_value))


def build_reports_export_text(overview):
    lines = [
        "# HospAI Reports Overview",
        "",
        "## Financial Summary",
        f"- Total billed: INR {overview['billing_summary']['total_billed']:.2f}",
        f"- Total collected: INR {overview['billing_summary']['total_collected']:.2f}",
        f"- Total due: INR {overview['billing_summary']['total_due']:.2f}",
        f"- Net position: INR {overview['accounts_summary']['net_position']:.2f}",
        "",
        "## Operations",
        f"- Monthly OP: {overview['hospital_summary']['ip_op_counts']['monthly_op']}",
        f"- Monthly IP: {overview['hospital_summary']['ip_op_counts']['monthly_ip']}",
        f"- Average LOS: {overview['alos_summary']['average_los_days']} days",
        f"- Admissions counted: {overview['alos_summary']['admission_count']}",
        "",
        "## Clinic Income",
    ]

    for row in overview.get("clinic_income", []):
        lines.append(f"- {row['label']}: INR {float(row['count']):.2f}")

    lines.extend(["", "## Discounts by Module"])
    for row in overview.get("discount_by_module", []):
        lines.append(f"- {row['label']}: INR {float(row['count']):.2f}")

    lines.extend(["", "## Payment Status"])
    for row in overview.get("payment_status_breakdown", []):
        lines.append(f"- {row['label']}: {row['count']} invoice(s)")

    return "\n".join(lines)




def format_inr_pdf(amount):
    try:
        value = float(amount or 0)
    except (TypeError, ValueError):
        value = 0.0
    return "â‚¹" + f"{value:,.0f}"


def _int_value(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _float_value(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _percent(part, whole):
    whole = _float_value(whole)
    if whole <= 0:
        return 0
    return round((_float_value(part) / whole) * 100)


def _short_pdf_text(value, limit=95):
    text = " ".join(str(value or "").split())
    return text if len(text) <= limit else text[: max(0, limit - 3)] + "..."


def _fetch_dashboard_pdf_data(hospital_id=None):
    """Build the dashboard print/export dataset from live HospAI tables.

    The PDF renderer keeps the executive dashboard layout stable, while this
    function keeps values dynamic. Queries are defensive so older offline
    databases that are missing optional modules still produce a report instead
    of failing the export.
    """
    scoped_hospital_id = hospital_id or current_hospital_id()
    total_beds = 250
    data = {
        "registrations_today": 0,
        "registrations_yesterday": 0,
        "op_waiting": 0,
        "op_in_consultation": 0,
        "op_completed": 0,
        "daily_op": 0,
        "daily_ip": 0,
        "today_revenue": 0.0,
        "monthly_revenue": 0.0,
        "lab_revenue": 0.0,
        "pharmacy_revenue": 0.0,
        "total_collected": 0.0,
        "outstanding_amount": 0.0,
        "occupied_beds": 0,
        "total_beds": total_beds,
        "pending_lab_reports": 0,
        "completed_lab_reports": 0,
        "lab_tests_today": 0,
        "pharmacy_pending_bills": 0,
        "pharmacy_bills_today": 0,
        "pharmacy_bills_completed": 0,
        "prescriptions_today": 0,
        "prescriptions_completed": 0,
        "followups_today": 0,
        "followups_completed": 0,
        "revenue_heads": {},
        "payment_modes": [],
        "low_stock_item": None,
        "low_stock_quantity": None,
    }

    def one(cursor, sql, params=(), key="value"):
        try:
            cursor.execute(sql, params)
            row = cursor.fetchone()
            return (row or {}).get(key, 0) if hasattr(row, "get") else (row[key] if row else 0)
        except Exception:
            return 0

    def rows(cursor, sql, params=()):
        try:
            cursor.execute(sql, params)
            return [dict(row) for row in cursor.fetchall()]
        except Exception:
            return []

    with get_connection() as conn:
        cursor = conn.cursor()
        data["registrations_today"] = _int_value(one(cursor, "SELECT COUNT(*) AS value FROM patients WHERE hospital_id = ? AND DATE(created_at) = CURRENT_DATE", (scoped_hospital_id,)))
        data["registrations_yesterday"] = _int_value(one(cursor, "SELECT COUNT(*) AS value FROM patients WHERE hospital_id = ? AND DATE(created_at) = DATE('now', '-1 day')", (scoped_hospital_id,)))
        data["daily_op"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM encounters e
            JOIN patients p ON p.patient_id = e.patient_id
            WHERE p.hospital_id = ? AND e.encounter_type = 'OP' AND DATE(e.arrival_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["daily_ip"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM encounters e
            JOIN patients p ON p.patient_id = e.patient_id
            WHERE p.hospital_id = ? AND e.encounter_type = 'IP' AND DATE(e.arrival_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["op_waiting"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM appointments a
            LEFT JOIN patients p ON p.patient_id = a.patient_id
            WHERE (p.hospital_id = ? OR a.patient_id IS NULL)
              AND DATE(a.appointment_date) = CURRENT_DATE
              AND LOWER(COALESCE(a.status, 'scheduled')) IN ('scheduled','waiting','queued','pending')
        """, (scoped_hospital_id,)))
        data["op_in_consultation"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM appointments a
            LEFT JOIN patients p ON p.patient_id = a.patient_id
            WHERE (p.hospital_id = ? OR a.patient_id IS NULL)
              AND DATE(a.appointment_date) = CURRENT_DATE
              AND LOWER(COALESCE(a.status, '')) IN ('in consultation','in_consultation','consultation','active')
        """, (scoped_hospital_id,)))
        data["op_completed"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM appointments a
            LEFT JOIN patients p ON p.patient_id = a.patient_id
            WHERE (p.hospital_id = ? OR a.patient_id IS NULL)
              AND DATE(a.appointment_date) = CURRENT_DATE
              AND LOWER(COALESCE(a.status, '')) IN ('completed','done','consulted','finished')
        """, (scoped_hospital_id,)))
        data["today_revenue"] = _float_value(one(cursor, """
            SELECT COALESCE(SUM(i.total_amount), 0) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND DATE(i.created_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["monthly_revenue"] = _float_value(one(cursor, """
            SELECT COALESCE(SUM(i.total_amount), 0) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND to_char(i.created_at, 'YYYY-MM')=to_char(CURRENT_DATE, 'YYYY-MM')
        """, (scoped_hospital_id,)))
        data["lab_revenue"] = _float_value(one(cursor, """
            SELECT COALESCE(SUM(d.amount), 0) AS value
            FROM diagnostics d
            LEFT JOIN patients p ON p.patient_id = d.patient_id
            WHERE (p.hospital_id = ? OR d.patient_id IS NULL) AND to_char(d.created_at, 'YYYY-MM')=to_char(CURRENT_DATE, 'YYYY-MM')
        """, (scoped_hospital_id,)))
        data["pharmacy_revenue"] = _float_value(one(cursor, """
            SELECT COALESCE(SUM(ps.amount), 0) AS value
            FROM pharmacy_sales ps
            LEFT JOIN patients p ON p.patient_id = ps.patient_id
            WHERE (p.hospital_id = ? OR ps.patient_id IS NULL) AND to_char(ps.sold_at, 'YYYY-MM')=to_char(CURRENT_DATE, 'YYYY-MM')
        """, (scoped_hospital_id,)))
        data["total_collected"] = _float_value(one(cursor, """
            SELECT COALESCE(SUM(ip.amount), 0) AS value
            FROM invoice_payments ip
            JOIN invoices i ON i.id = ip.invoice_id
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND DATE(ip.created_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["outstanding_amount"] = _float_value(one(cursor, """
            SELECT COALESCE(SUM(i.due_amount), 0) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND i.due_amount > 0
        """, (scoped_hospital_id,)))
        data["occupied_beds"] = _int_value(one(cursor, "SELECT COUNT(*) AS value FROM admissions WHERE hospital_id = ? AND discharge_date IS NULL", (scoped_hospital_id,)))
        data["lab_tests_today"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM diagnostics d
            LEFT JOIN patients p ON p.patient_id = d.patient_id
            WHERE (p.hospital_id = ? OR d.patient_id IS NULL) AND DATE(d.created_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["completed_lab_reports"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM diagnostics d
            LEFT JOIN patients p ON p.patient_id = d.patient_id
            WHERE (p.hospital_id = ? OR d.patient_id IS NULL)
              AND DATE(d.created_at) = CURRENT_DATE
              AND (LOWER(COALESCE(d.order_status,'')) IN ('reported','completed','verified') OR d.reported_at IS NOT NULL)
        """, (scoped_hospital_id,)))
        data["pending_lab_reports"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM diagnostics d
            LEFT JOIN patients p ON p.patient_id = d.patient_id
            WHERE (p.hospital_id = ? OR d.patient_id IS NULL)
              AND LOWER(COALESCE(d.order_status,'ordered')) NOT IN ('reported','completed','verified','cancelled')
        """, (scoped_hospital_id,)))
        data["pharmacy_pending_bills"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL)
              AND LOWER(COALESCE(i.module,'')) LIKE '%pharm%'
              AND COALESCE(i.due_amount, 0) > 0
        """, (scoped_hospital_id,)))
        data["pharmacy_bills_today"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL)
              AND LOWER(COALESCE(i.module,'')) LIKE '%pharm%'
              AND DATE(i.created_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["pharmacy_bills_completed"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL)
              AND LOWER(COALESCE(i.module,'')) LIKE '%pharm%'
              AND DATE(i.created_at) = CURRENT_DATE
              AND LOWER(COALESCE(i.payment_status,'')) IN ('paid','settled','completed')
        """, (scoped_hospital_id,)))
        data["prescriptions_today"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM medication_schedules m
            LEFT JOIN patients p ON p.patient_id = m.patient_id
            WHERE (p.hospital_id = ? OR m.patient_id IS NULL) AND DATE(m.created_at) = CURRENT_DATE
        """, (scoped_hospital_id,)))
        data["prescriptions_completed"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM medication_schedules m
            LEFT JOIN patients p ON p.patient_id = m.patient_id
            WHERE (p.hospital_id = ? OR m.patient_id IS NULL)
              AND DATE(m.created_at) = CURRENT_DATE
              AND LOWER(COALESCE(m.status,'')) IN ('completed','dispensed','done')
        """, (scoped_hospital_id,)))
        data["followups_today"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM appointments a
            LEFT JOIN patients p ON p.patient_id = a.patient_id
            WHERE (p.hospital_id = ? OR a.patient_id IS NULL)
              AND DATE(a.appointment_date) = CURRENT_DATE
              AND (LOWER(COALESCE(a.appointment_kind,'')) = 'follow_up' OR a.follow_up_for IS NOT NULL)
        """, (scoped_hospital_id,)))
        data["followups_completed"] = _int_value(one(cursor, """
            SELECT COUNT(*) AS value
            FROM appointments a
            LEFT JOIN patients p ON p.patient_id = a.patient_id
            WHERE (p.hospital_id = ? OR a.patient_id IS NULL)
              AND DATE(a.appointment_date) = CURRENT_DATE
              AND (LOWER(COALESCE(a.appointment_kind,'')) = 'follow_up' OR a.follow_up_for IS NOT NULL)
              AND LOWER(COALESCE(a.status,'')) IN ('completed','done','consulted','finished')
        """, (scoped_hospital_id,)))
        revenue_rows = rows(cursor, """
            SELECT LOWER(COALESCE(NULLIF(TRIM(i.module), ''), 'other')) AS label,
                   COALESCE(SUM(i.total_amount), 0) AS count
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND DATE(i.created_at) = CURRENT_DATE
            GROUP BY label
        """, (scoped_hospital_id,))
        for row in revenue_rows:
            label = row.get("label") or "other"
            amount = _float_value(row.get("count"))
            if "lab" in label or "diagnostic" in label:
                data["revenue_heads"]["Lab Revenue"] = data["revenue_heads"].get("Lab Revenue", 0) + amount
            elif "pharm" in label:
                data["revenue_heads"]["Pharmacy Revenue"] = data["revenue_heads"].get("Pharmacy Revenue", 0) + amount
            elif "ip" in label:
                data["revenue_heads"]["IPD Revenue"] = data["revenue_heads"].get("IPD Revenue", 0) + amount
            elif "op" in label or "consult" in label:
                data["revenue_heads"]["OPD Revenue"] = data["revenue_heads"].get("OPD Revenue", 0) + amount
            else:
                data["revenue_heads"]["Other Revenue"] = data["revenue_heads"].get("Other Revenue", 0) + amount
        data["payment_modes"] = rows(cursor, """
            SELECT COALESCE(NULLIF(TRIM(ip.payment_mode), ''), 'Unknown') AS label,
                   COALESCE(SUM(ip.amount), 0) AS count
            FROM invoice_payments ip
            JOIN invoices i ON i.id = ip.invoice_id
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND DATE(ip.created_at) = CURRENT_DATE
            GROUP BY label
            ORDER BY count DESC, label ASC
        """, (scoped_hospital_id,))
        low_stock = rows(cursor, """
            SELECT medicine_name AS label, quantity AS count
            FROM pharmacy_purchases
            WHERE COALESCE(stock_applied, 0) = 1 AND COALESCE(quantity, 0) <= 50
            ORDER BY quantity ASC, medicine_name ASC
            LIMIT 1
        """)
        if low_stock:
            data["low_stock_item"] = low_stock[0].get("label")
            data["low_stock_quantity"] = _int_value(low_stock[0].get("count"))
        data["due_patients_count"] = _int_value(one(cursor, """
            SELECT COUNT(DISTINCT i.patient_id) AS value
            FROM invoices i
            LEFT JOIN patients p ON p.patient_id = i.patient_id
            WHERE (p.hospital_id = ? OR i.patient_id IS NULL) AND i.due_amount > 10000
        """, (scoped_hospital_id,)))
    return data




def _dashboard_report_logo_path():
    """Return the best available hospital logo for dashboard PDF headers.

    Keep this defensive because the offline package can be started either from
    the project root or the backend folder. The PDF must still generate if the
    logo file is moved or missing.

    Search order:
      1. backend/assets/  (highest priority â€“ always present beside app.py)
      2. assets/ at project root
      3. frontend/public/ and frontend/dist/
      4. PyInstaller _MEIPASS (packaged build)
      5. Exe-beside directory (desktop packaged release)
      6. CWD fallbacks
    """
    import sys as _sys

    # Directory that contains app.py (i.e. backend/)
    backend_dir = os.path.dirname(os.path.abspath(__file__))
    # Project root is one level above backend/
    project_root = os.path.abspath(os.path.join(backend_dir, ".."))
    meipass = getattr(_sys, "_MEIPASS", "")
    # Directory beside the packaged exe (e.g. HospAI_Backend.exe)
    exe_dir = os.path.dirname(os.path.abspath(_sys.executable)) if getattr(_sys, "frozen", False) else ""

    candidates = [
        # 1. backend/assets/ â€” present after build and during dev
        os.path.join(backend_dir, "assets", "rapha_logo.png"),
        os.path.join(backend_dir, "assets", "rapha_logo.jpg"),
        # 2. Project root assets/
        os.path.join(project_root, "assets", "rapha_logo.png"),
        os.path.join(project_root, "assets", "rapha_logo.jpg"),
        # 3. Frontend public/dist
        os.path.join(project_root, "frontend", "public", "rapha_logo.png"),
        os.path.join(project_root, "frontend", "public", "rapha_logo.jpg"),
        os.path.join(project_root, "frontend", "public", "logo.png"),
        os.path.join(project_root, "frontend", "public", "logo.jpg"),
        os.path.join(project_root, "frontend", "public", "logo_square.png"),
        os.path.join(project_root, "frontend", "dist", "rapha_logo.png"),
        os.path.join(project_root, "frontend", "dist", "rapha_logo.jpg"),
        os.path.join(project_root, "frontend", "dist", "logo.png"),
        os.path.join(project_root, "frontend", "dist", "logo.jpg"),
        os.path.join(project_root, "frontend", "dist", "logo_square.png"),
        # 4. PyInstaller _MEIPASS (flat and in assets/ sub-dir)
        os.path.join(meipass, "assets", "rapha_logo.png") if meipass else "",
        os.path.join(meipass, "assets", "rapha_logo.jpg") if meipass else "",
        os.path.join(meipass, "rapha_logo.png") if meipass else "",
        os.path.join(meipass, "rapha_logo.jpg") if meipass else "",
        # 5. Beside the packaged exe
        os.path.join(exe_dir, "assets", "rapha_logo.png") if exe_dir else "",
        os.path.join(exe_dir, "assets", "rapha_logo.jpg") if exe_dir else "",
        os.path.join(exe_dir, "rapha_logo.png") if exe_dir else "",
        os.path.join(exe_dir, "rapha_logo.jpg") if exe_dir else "",
        # 6. CWD fallbacks
        os.path.join(os.getcwd(), "assets", "rapha_logo.png"),
        os.path.join(os.getcwd(), "assets", "rapha_logo.jpg"),
        os.path.join(project_root, "assets", "rapha_print_header.jpg"),
        os.path.join(project_root, "frontend", "public", "rapha_print_header.jpg"),
        os.path.join(project_root, "logo.png"),
        os.path.join(project_root, "assets", "hosp_ai_logo.png"),
        os.path.join(project_root, "assets", "logo.jpg"),
        os.path.join(os.getcwd(), "assets", "rapha_print_header.jpg"),
        os.path.join(os.getcwd(), "logo.png"),
        os.path.join(os.getcwd(), "assets", "hosp_ai_logo.png"),
    ]

    for candidate in candidates:
        if candidate and os.path.exists(candidate) and os.path.getsize(candidate) > 0:
            return candidate

    # No logo found â€” write a log entry so the operator knows
    try:
        log_dir = os.path.join(os.environ.get("APPDATA", ""), "HospAI", "logs")
        os.makedirs(log_dir, exist_ok=True)
        log_path = os.path.join(log_dir, "print_logo.log")
        with open(log_path, "a", encoding="utf-8") as _lf:
            import datetime as _dt
            _lf.write(
                f"[{_dt.datetime.now().isoformat()}] WARNING: rapha_logo not found. "
                f"backend_dir={backend_dir!r}, project_root={project_root!r}, meipass={meipass!r}\n"
            )
    except Exception:
        pass

    return None

def format_header_date(dt):
    # Format: month/day/short_year, Hour:Minute AM/PM (uppercase)
    month = dt.month
    day = dt.day
    short_year = dt.strftime("%y")
    time_str = dt.strftime("%I:%M %p")
    if time_str.startswith("0"):
        time_str = time_str[1:]
    return f"{month}/{day}/{short_year}, {time_str}"


def format_print_date(dt):
    # Format: day/month/year, Hour:Minute:Second pm/am (lowercase)
    day = dt.day
    month = dt.month
    year = dt.year
    time_str = dt.strftime("%I:%M:%S %p").lower()
    return f"{day}/{month}/{year}, {time_str}"


class DashboardPDF(FPDF):
    def __init__(self, font_name="Helvetica", print_now=None, hostname="localhost:5001"):
        super().__init__(orientation="P", unit="mm", format="A4")
        self.font_name = font_name
        self.print_now = print_now
        self.hostname = hostname

    def header(self):
        # Continuation pages only: show a compact running header with date/time.
        # Page 1 gets a full branding header drawn manually in generate_executive_dashboard_pdf.
        if self.page_no() == 1:
            return  # Page 1 header is drawn explicitly; skip auto-header.
        self.set_y(5)
        self.set_font(self.font_name, "", 7)
        self.set_text_color(80, 80, 80)
        # Left: date/time
        header_time = format_header_date(self.print_now)
        self.set_x(15)
        self.cell(60, 5, header_time, align="L")
        # Center: hospital name (compact, no overlap)
        self.set_x(75)
        self.cell(60, 5, "VERARA POLYCLINIC, PHARMACY, DIAGNOSTICS", align="C")
        # Right: page label
        self.set_x(135)
        self.cell(60, 5, f"Page {self.page_no()}/{{nb}}", align="R")
        self.set_text_color(0, 0, 0)

    def footer(self):
        # Bottom-left and bottom-right footer on every page
        self.set_y(-10) # 10mm from bottom
        self.set_font(self.font_name, "", 7)
        self.set_text_color(0, 0, 0)
        
        # Left-aligned Hostname
        self.set_x(15)
        self.cell(90, 5, self.hostname, align="L")
        
        # Right-aligned Page Number
        self.set_x(105)
        self.cell(90, 5, f"{self.page_no()}/{{nb}}", align="R")


def generate_executive_dashboard_pdf(hospital_id=None, hostname="localhost:5001"):
    now = current_ist_datetime()
    data = _fetch_dashboard_pdf_data(hospital_id=hospital_id)
    total_beds = max(_int_value(data["total_beds"]), 1)
    occupied_beds = min(_int_value(data["occupied_beds"]), total_beds)
    available_beds = max(total_beds - occupied_beds, 0)
    bed_occupancy_pct = _percent(occupied_beds, total_beds)
    queue_total = _int_value(data["op_waiting"]) + _int_value(data["op_in_consultation"]) + _int_value(data["op_completed"])
    registration_delta = _int_value(data["registrations_today"]) - _int_value(data["registrations_yesterday"])
    registration_delta_pct = _percent(registration_delta, max(_int_value(data["registrations_yesterday"]), 1)) if data["registrations_yesterday"] else 0

    # Search for TrueType Unicode fonts to support Rupee symbol
    font = "Helvetica"
    
    # Initialize a temporary PDF to test font adding
    temp_pdf = FPDF()
    candidates = [
        # Linux standard DejaVu Sans
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "DejaVu"),
        # Windows standard Arial paths
        ("C:\\Windows\\Fonts\\arial.ttf", "C:\\Windows\\Fonts\\arialbd.ttf", "Arial"),
        ("C:\\Windows\\Fonts\\calibri.ttf", "C:\\Windows\\Fonts\\calibrib.ttf", "Calibri"),
    ]
    
    loaded_font_info = None
    for reg, bld, name in candidates:
        if os.path.exists(reg) and os.path.exists(bld):
            try:
                temp_pdf.add_font(name, "", reg)
                temp_pdf.add_font(name, "B", bld)
                font = name
                loaded_font_info = (reg, bld, name)
                break
            except Exception:
                pass

    pdf = DashboardPDF(font_name=font, print_now=now, hostname=hostname)
    pdf.set_left_margin(15)
    pdf.set_right_margin(15)
    pdf.set_top_margin(15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.alias_nb_pages()
    
    # Load fonts into the actual PDF if available
    if loaded_font_info:
        try:
            pdf.add_font(loaded_font_info[2], "", loaded_font_info[0])
            pdf.add_font(loaded_font_info[2], "B", loaded_font_info[1])
        except Exception:
            pdf.font_name = "Helvetica"
            font = "Helvetica"

    pdf.add_page()

    def clean_txt(text):
        if not text:
            return ""
        text = str(text)
        if pdf.font_name in ["Helvetica", "Times-Roman", "Courier", "Times"]:
            text = text.replace("â‚¹", "Rs. ")
            return text.encode("latin-1", errors="replace").decode("latin-1")
        return text

    def ensure_space(height):
        # Trigger page break if space is insufficient
        if pdf.get_y() + height > 282:
            pdf.add_page()
            # The add_page resets self.y to top_margin (15).
            pdf.set_y(15)

    def section(title):
        ensure_space(15)
        pdf.set_font(font, "B", 9.5)
        pdf.cell(180, 6, clean_txt(title), ln=True)
        current_y = pdf.get_y()
        pdf.set_line_width(0.3)
        pdf.line(15, current_y, 195, current_y)
        pdf.ln(2)

    def table(headers, rows_, widths):
        row_h = 6.4
        ensure_space(row_h * 2) # Ensure space for header + at least 1 row
        
        # Table Header
        pdf.set_font(font, "B", 7.5)
        for h, w in zip(headers, widths):
            pdf.cell(w, row_h, clean_txt(h), border=1, align="L")
        pdf.ln(row_h)
        
        # Table Data
        pdf.set_font(font, "", 7.5)
        for row in rows_:
            ensure_space(row_h)
            for val, w in zip(row, widths):
                pdf.cell(w, row_h, clean_txt(val), border=1, align="L")
            pdf.ln(row_h)
        pdf.ln(3)

    # Page 1 Header Block â€” proper 3-column layout:
    # Left col:  Logo (x=15, y=14, w=22mm)
    # Center col: Hospital details (x=42, y=14)
    # Right col:  Report title (x=115, y=16, right-aligned)
    # Separator line at y=42mm; first table starts at y=48mm
    LOGO_X = 15
    LOGO_Y = 14
    LOGO_W = 42
    TEXT_X = 61
    TITLE_X = 115
    SEP_Y = 42

    logo_path = _dashboard_report_logo_path()
    if logo_path:
        try:
            pdf.image(logo_path, x=LOGO_X, y=LOGO_Y, w=LOGO_W)
        except Exception:
            logo_path = None

    # Hospital details block (center-left)
    if not logo_path:
        TEXT_X = LOGO_X  # shift left if no logo
    pdf.set_xy(TEXT_X, LOGO_Y)
    pdf.set_font(font, "B", 16)
    pdf.set_text_color(20, 61, 94)
    pdf.cell(70, 6, clean_txt("VERARA"), ln=True)
    pdf.set_x(TEXT_X)
    pdf.set_font(font, "B", 10)
    pdf.set_text_color(20, 61, 94)
    pdf.cell(70, 5, clean_txt("POLYCLINIC, PHARMACY,"), ln=True)
    pdf.set_x(TEXT_X)
    pdf.cell(70, 5, clean_txt("DIAGNOSTICS"), ln=True)
    pdf.set_text_color(0, 0, 0)

    # Report title block (right side)
    pdf.set_xy(TITLE_X, LOGO_Y + 2)
    pdf.set_font(font, "B", 14)
    pdf.set_text_color(15, 45, 76)
    pdf.cell(80, 7, clean_txt("Executive Dashboard Report"), ln=True, align="R")
    pdf.set_x(TITLE_X)
    pdf.set_font(font, "", 8)
    pdf.set_text_color(83, 101, 120)
    pdf.cell(80, 5, clean_txt("VERARA hospital performance & revenue"), ln=True, align="R")
    pdf.set_x(TITLE_X)
    pdf.cell(80, 5, clean_txt(f"Print Date: {format_print_date(now)}"), ln=True, align="R")
    pdf.set_text_color(0, 0, 0)

    # Separator line â€” drawn at y=42mm, safely below all header content
    pdf.set_draw_color(13, 148, 169)
    pdf.set_line_width(0.6)
    pdf.line(15, SEP_Y, 195, SEP_Y)
    # First table starts at y=48mm
    pdf.set_y(48)

    # 1. Dashboard Summary
    section("Dashboard Summary")
    table(
        ["Metric", "Value", "Remarks"],
        [
            ["Today's Registrations", str(data["registrations_today"]), "New patient registrations today"],
            ["Today's Revenue", format_inr_pdf(data["today_revenue"]), "Live invoice total for today"],
            ["Monthly Revenue", format_inr_pdf(data["monthly_revenue"]), "Live invoice total for this month"],
            ["Lab Revenue", format_inr_pdf(data["lab_revenue"]), "Current month diagnostics revenue"],
            ["Pharmacy Revenue", format_inr_pdf(data["pharmacy_revenue"]), "Current month pharmacy sales"],
        ],
        [65, 35, 80],
    )

    # 2. Today's Operations
    section("Today's Operations")
    table(
        ["Module", "Count", "Completed", "Pending"],
        [
            ["OP Patients", str(data["daily_op"] or queue_total), str(data["op_completed"]), str(data["op_waiting"])],
            ["IP Patients", str(data["daily_ip"]), str(max(_int_value(data["daily_ip"]) - occupied_beds, 0)), str(occupied_beds)],
            ["Lab Tests", str(data["lab_tests_today"]), str(data["completed_lab_reports"]), str(data["pending_lab_reports"])],
            ["Lab Reports", str(data["lab_tests_today"]), str(data["completed_lab_reports"]), str(data["pending_lab_reports"])],
            ["Pharmacy Bills", str(data["pharmacy_bills_today"]), str(data["pharmacy_bills_completed"]), str(data["pharmacy_pending_bills"])],
            ["Prescriptions", str(data["prescriptions_today"]), str(data["prescriptions_completed"]), str(max(_int_value(data["prescriptions_today"]) - _int_value(data["prescriptions_completed"]), 0))],
            ["Follow Ups", str(data["followups_today"]), str(data["followups_completed"]), str(max(_int_value(data["followups_today"]) - _int_value(data["followups_completed"]), 0))],
        ],
        [50, 40, 45, 45],
    )

    # 3. Revenue Snapshot
    revenue_heads = data["revenue_heads"]
    op_rev = revenue_heads.get("OPD Revenue", 0)
    lab_rev = revenue_heads.get("Lab Revenue", 0)
    pharm_rev = revenue_heads.get("Pharmacy Revenue", 0)
    ip_rev = revenue_heads.get("IPD Revenue", 0)
    other_rev = revenue_heads.get("Other Revenue", max(_float_value(data["today_revenue"]) - op_rev - lab_rev - pharm_rev - ip_rev, 0))
    
    section("Revenue Snapshot")
    table(
        ["Revenue Head", "Amount"],
        [
            ["Today Revenue", format_inr_pdf(data["today_revenue"])],
            ["Monthly Revenue", format_inr_pdf(data["monthly_revenue"])],
            ["Lab Revenue", format_inr_pdf(data["lab_revenue"])],
            ["Pharmacy Revenue", format_inr_pdf(data["pharmacy_revenue"])],
            ["Outstanding Amount", format_inr_pdf(data["outstanding_amount"])],
        ],
        [120, 60],
    )

    # 4. Payment Summary
    payment_rows = [[row.get("label") or "Unknown", format_inr_pdf(row.get("count"))] for row in data["payment_modes"]]
    if not payment_rows:
        payment_rows = [["No collections recorded today", format_inr_pdf(0)]]
    payment_rows.append(["Total Collected", format_inr_pdf(data["total_collected"])])
    
    section("Payment Summary")
    table(
        ["Payment Mode", "Amount"],
        payment_rows,
        [120, 60],
    )

    data_bytes = pdf.output(dest="S")
    if isinstance(data_bytes, str):
        return data_bytes.encode("latin-1")
    return bytes(data_bytes)


@app.post("/api/share/upload")
@require_session
def share_upload():
    import uuid
    if "file" not in request.files:
        return jsonify({"error": "Missing file"}), 400
    uploaded_file = request.files["file"]
    file_bytes = uploaded_file.read()
    
    filename = secure_filename(uploaded_file.filename or "export.bin")
    mimetype = uploaded_file.mimetype or "application/octet-stream"
    file_path = STORAGE.store(
        doc_type="shared_exports",
        file_name=filename,
        data=file_bytes,
        mime_type=mimetype
    )
    
    share_token = uuid.uuid4().hex
    create_shared_export(share_token, file_path, filename, mimetype)
    
    share_url = f"{request.host_url}api/share/view/{share_token}"
    return jsonify({"share_url": share_url, "share_token": share_token})


@app.get("/api/share/view/<share_token>")
def share_view(share_token):
    record = get_shared_export(share_token)
    if not record:
        return jsonify({"error": "Shared export not found"}), 404
        
    file_path = record["file_path"]
    file_name = record["file_name"] or "export"
    mime_type = record["mime_type"] or "application/octet-stream"
    
    file_bytes = STORAGE.read(file_path)
    if not file_bytes:
        return jsonify({"error": "Shared file content unavailable"}), 404
        
    response = send_file(
        io.BytesIO(file_bytes),
        mimetype=mime_type,
        as_attachment=False,
        download_name=file_name
    )
    response.headers["Cache-Control"] = "public, max-age=86400"
    return response


def _build_dashboard_export_payload():
    summary = get_hospital_dashboard_summary()
    revenue = summary.get("revenue") or {}
    payment_summary = summary.get("payment_summary") or {}
    return {
        "generated_at": current_ist_datetime().strftime("%Y-%m-%d %H:%M:%S"),
        "revenue": {
            "today": revenue.get("today_total", 0) or 0,
            "weekly": revenue.get("weekly_revenue", revenue.get("weekly_total", 0)) or 0,
            "monthly": revenue.get("monthly_total", 0) or 0,
            "yearly": revenue.get("yearly_total", 0) or 0,
        },
        "payment_summary": {
            "total_collection": payment_summary.get("total_collection", revenue.get("total_collection", 0)) or 0,
            "pending_payments": payment_summary.get("pending_payments", revenue.get("pending_payments", 0)) or 0,
            "paid_payments": payment_summary.get("paid_payments", revenue.get("paid_payments", 0)) or 0,
            "today_collection": payment_summary.get("today_collection", revenue.get("today_collection", 0)) or 0,
        },
        "operations_today": summary.get("operations_today") or {},
        "ip_op_counts": summary.get("ip_op_counts") or {},
    }


@app.get("/api/dashboard/export/pdf")
@require_session
def dashboard_export_pdf():
    try:
        pdf_bytes = generate_executive_dashboard_pdf(hospital_id=current_hospital_id(), hostname=(request.host or "localhost:5001"))
        response = send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name="executive-dashboard.pdf",
        )
        response.headers["Content-Disposition"] = 'attachment; filename="executive-dashboard.pdf"'
        response.headers["Access-Control-Expose-Headers"] = "Content-Disposition, Content-Length, Content-Type"
        response.headers["Cache-Control"] = "no-store, max-age=0"
        return response
    except Exception as exc:
        return jsonify({"error": "Unable to generate dashboard PDF.", "message": str(exc)}), 500


@app.get("/api/dashboard/export/csv")
@require_session
def dashboard_export_csv():
    try:
        payload = _build_dashboard_export_payload()
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["metric", "value"])
        writer.writerow(["today_revenue", payload["revenue"]["today"]])
        writer.writerow(["weekly_revenue", payload["revenue"]["weekly"]])
        writer.writerow(["monthly_revenue", payload["revenue"]["monthly"]])
        writer.writerow(["yearly_revenue", payload["revenue"]["yearly"]])
        writer.writerow(["total_collection", payload["payment_summary"]["total_collection"]])
        writer.writerow(["pending_payments", payload["payment_summary"]["pending_payments"]])
        writer.writerow(["paid_payments", payload["payment_summary"]["paid_payments"]])
        writer.writerow(["today_collection", payload["payment_summary"]["today_collection"]])
        csv_bytes = buffer.getvalue().encode("utf-8")
        return send_file(
            io.BytesIO(csv_bytes),
            mimetype="text/csv",
            as_attachment=True,
            download_name="executive-dashboard.csv",
        )
    except Exception as exc:
        return jsonify({"error": "Unable to generate dashboard CSV.", "message": str(exc)}), 500


@app.get("/api/dashboard/export/excel")
@require_session
def dashboard_export_excel():
    try:
        from openpyxl import Workbook

        payload = _build_dashboard_export_payload()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dashboard"
        sheet.append(["metric", "value"])
        sheet.append(["today_revenue", payload["revenue"]["today"]])
        sheet.append(["weekly_revenue", payload["revenue"]["weekly"]])
        sheet.append(["monthly_revenue", payload["revenue"]["monthly"]])
        sheet.append(["yearly_revenue", payload["revenue"]["yearly"]])
        sheet.append(["total_collection", payload["payment_summary"]["total_collection"]])
        sheet.append(["pending_payments", payload["payment_summary"]["pending_payments"]])
        sheet.append(["paid_payments", payload["payment_summary"]["paid_payments"]])
        sheet.append(["today_collection", payload["payment_summary"]["today_collection"]])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="executive-dashboard.xlsx",
        )
    except Exception as exc:
        return jsonify({"error": "Unable to generate dashboard Excel export.", "message": str(exc)}), 500


@app.get("/api/dashboard/print/pdf")
@require_session
def dashboard_print_pdf():
    # Backward-compatible print endpoint for clients that call a dedicated print URL.
    return dashboard_export_pdf()


@app.get("/api/reports/export/csv")
@require_permissions("reports.read")
def reports_export_csv():
    overview = get_reports_overview()
    csv_stream = io.StringIO()
    writer = csv.writer(csv_stream)
    writer.writerow(["section", "label", "value"])
    writer.writerow(["billing", "total_billed", overview["billing_summary"]["total_billed"]])
    writer.writerow(["billing", "total_collected", overview["billing_summary"]["total_collected"]])
    writer.writerow(["billing", "total_due", overview["billing_summary"]["total_due"]])
    writer.writerow(["accounts", "net_position", overview["accounts_summary"]["net_position"]])
    writer.writerow(["operations", "monthly_op", overview["hospital_summary"]["ip_op_counts"]["monthly_op"]])
    writer.writerow(["operations", "monthly_ip", overview["hospital_summary"]["ip_op_counts"]["monthly_ip"]])
    writer.writerow(["operations", "average_los_days", overview["alos_summary"]["average_los_days"]])
    for row in overview.get("clinic_income", []):
        writer.writerow(["clinic_income", row["label"], row["count"]])
    for row in overview.get("discount_by_module", []):
        writer.writerow(["discount_by_module", row["label"], row["count"]])
    for row in overview.get("payment_status_breakdown", []):
        writer.writerow(["payment_status", row["label"], row["count"]])

    payload = io.BytesIO(csv_stream.getvalue().encode("utf-8"))
    payload.seek(0)
    return send_file(
        payload,
        mimetype="text/csv",
        as_attachment=True,
        download_name="reports-overview.csv",
    )


@app.get("/api/reports/export/pdf")
@require_permissions("reports.read")
def reports_export_pdf():
    overview = get_reports_overview()
    pdf_bytes = generate_pdf(
        "Hospital Operations",
        "reports_overview",
        build_reports_export_text(overview),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="reports-overview.pdf",
    )


@app.get("/api/reports/export/word")
@require_permissions("reports.read")
def reports_export_word():
    overview = get_reports_overview()
    word_bytes = generate_word(
        "Hospital Operations",
        "reports_overview",
        build_reports_export_text(overview),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return send_file(
        io.BytesIO(word_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        as_attachment=True,
        download_name="reports-overview.docx",
    )


@app.get("/api/stats")
@require_permissions("patients.read")
def stats():
    return jsonify(get_patient_stats(hospital_id=current_hospital_id()))


@app.get("/api/dashboard/analytics")
@require_permissions("patients.read")
def dashboard_analytics():
    requested_days = request.args.get("days", default=14, type=int)
    permissions = resolve_user_permissions(g.current_user)
    include_employee = "admin.use" in permissions
    try:
        payload = get_dashboard_analytics(
            days=requested_days,
            include_employee=include_employee,
            hospital_id=current_hospital_id(),
        )
    except TypeError as error:
        # Backward compatibility for stale runtime modules that still expose
        # get_dashboard_analytics(days, include_employee) without hospital_id.
        if "unexpected keyword argument 'hospital_id'" not in str(error):
            raise
        payload = get_dashboard_analytics(
            days=requested_days,
            include_employee=include_employee,
        )
    return jsonify(payload)


@app.get("/api/dashboard/hospital-summary")
@require_permissions("patients.read")
def dashboard_hospital_summary():
    selected_date = request.args.get("date")
    return jsonify(get_hospital_dashboard_summary(selected_date=selected_date))


@app.get("/api/patients")
@require_permissions("patients.read")
def patients_list():
    query = (request.args.get("q") or "").strip()
    hospital_id = current_hospital_id()
    patients = (
        search_patients(query, hospital_id=hospital_id)
        if query
        else get_all_patients(hospital_id=hospital_id)
    )
    return jsonify({"patients": rows_to_dicts(patients)})


@app.get("/api/appointments")
@require_permissions("patients.read")
def appointments_list():
    appointment_date = request.args.get("date")
    status = request.args.get("status")
    visit_type = request.args.get("visit_type")
    doctor_name = request.args.get("doctor_name")
    patient_id = request.args.get("patient_id")
    return jsonify(
        {
            "appointments": rows_to_dicts(
                list_appointments(
                    appointment_date=appointment_date,
                    status=status,
                    visit_type=visit_type,
                    doctor_name=doctor_name,
                    patient_id=patient_id,
                )
            )
        }
    )


@app.get("/api/doctors-history")
@require_permissions("patients.read")
def doctors_history_list():
    doctor_name = request.args.get("doctor_name")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    department = request.args.get("department")
    hospital_id = current_hospital_id()
    
    rows = list_doctors_history(
        hospital_id=hospital_id,
        doctor_name=doctor_name,
        from_date=from_date,
        to_date=to_date,
        department=department
    )
    return jsonify({"history": rows_to_dicts(rows)})


@app.post("/api/appointments")
@require_permissions("patients.write")
def appointments_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload, ["patient_name", "visit_type", "appointment_date", "consultation_fee"]
    )
    if validation_error:
        return validation_error
    try:
        consultation_fee = float(payload.get("consultation_fee") or payload.get("amount") or 0)
    except (TypeError, ValueError):
        consultation_fee = 0
    if consultation_fee <= 0:
        return jsonify({"error": "Consultation Fee is mandatory and must be greater than 0."}), 400
    appointment_id, token_no = create_appointment(payload)
    invoice_id = None
    payment_id = None
    if consultation_fee > 0:
        invoice_no = f"INV-OP-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        invoice_id = create_invoice(
            {
                "invoice_no": invoice_no,
                "patient_id": payload.get("patient_id"),
                "module": "OP",
                "doctor_name": payload.get("doctor_name"),
                "total_amount": consultation_fee,
                "paid_amount": 0,
                "advance_amount": 0,
                "refunded_amount": 0,
                "payment_status": "due",
                "created_by": g.current_user.get("username"),
            }
        )
        payment_id = record_invoice_payment(
            invoice_id,
            {
                "amount": consultation_fee,
                "payment_mode": normalize_payment_mode(payload.get("payment_mode") or "cash"),
                "gateway_ref": payload.get("gateway_ref"),
            },
        )
    log_audit_event(
        "create",
        "appointments",
        str(appointment_id),
        {"patient_name": payload.get("patient_name"), "token_no": token_no, "invoice_id": invoice_id, "payment_id": payment_id},
    )
    return jsonify({"appointment_id": appointment_id, "token_no": token_no, "invoice_id": invoice_id, "payment_id": payment_id})


@app.put("/api/appointments/<int:appointment_id>")
@require_permissions("patients.write")
def appointments_update(appointment_id):
    payload = request.get_json(force=True)
    updated = update_appointment(appointment_id, payload)
    if not updated:
        return jsonify({"error": "Appointment not found"}), 404
    log_audit_event(
        "update",
        "appointments",
        str(appointment_id),
        {"status": payload.get("status")},
    )
    return jsonify({"status": "ok"})


@app.post("/api/appointments/razorpay/order")
@require_permissions("patients.write")
def appointments_razorpay_order():
    config_error = require_razorpay_configured()
    if config_error:
        return config_error
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["amount"])
    if validation_error:
        return validation_error

    amount_paise = to_amount_paise(payload.get("amount"))
    if amount_paise is None:
        return jsonify({"error": "Amount must be greater than zero."}), 400

    receipt = payload.get("receipt") or f"appt-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    notes = payload.get("notes") if isinstance(payload.get("notes"), dict) else {}
    notes = {
        **notes,
        "hospital_id": str(current_hospital_id()),
        "created_by": g.current_user.get("username") or "",
    }
    try:
        order = create_razorpay_order(
            amount_paise=amount_paise,
            currency=payload.get("currency", "INR"),
            receipt=receipt,
            notes=notes,
        )
    except RuntimeError as error:
        return jsonify({"error": str(error)}), 502

    return jsonify(
        {
            "key_id": RAZORPAY_KEY_ID,
            "order_id": order.get("id"),
            "amount": order.get("amount"),
            "currency": order.get("currency"),
            "receipt": order.get("receipt"),
        }
    )


@app.post("/api/appointments/razorpay/verify")
@require_permissions("patients.write")
def appointments_razorpay_verify():
    config_error = require_razorpay_configured()
    if config_error:
        return config_error
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload,
        [
            "razorpay_order_id",
            "razorpay_payment_id",
            "razorpay_signature",
            "amount",
            "appointment",
        ],
    )
    if validation_error:
        return validation_error

    if not verify_razorpay_signature(
        payload.get("razorpay_order_id"),
        payload.get("razorpay_payment_id"),
        payload.get("razorpay_signature"),
    ):
        return jsonify({"error": "Invalid Razorpay signature."}), 400

    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than zero."}), 400

    appointment_payload = payload.get("appointment")
    if not isinstance(appointment_payload, dict):
        return jsonify({"error": "appointment must be an object."}), 400

    appointment_validation = validate_required_fields(
        appointment_payload, ["patient_name", "visit_type", "appointment_date"]
    )
    if appointment_validation:
        return appointment_validation

    appointment_id, token_no = create_appointment(
        {
            "patient_id": appointment_payload.get("patient_id"),
            "patient_name": appointment_payload.get("patient_name"),
            "visit_type": appointment_payload.get("visit_type", "OP"),
            "department": appointment_payload.get("department"),
            "doctor_name": appointment_payload.get("doctor_name"),
            "appointment_date": appointment_payload.get("appointment_date"),
            "status": appointment_payload.get("status", "scheduled"),
            "appointment_kind": appointment_payload.get("appointment_kind", "new"),
            "follow_up_for": appointment_payload.get("follow_up_for"),
            "notes": appointment_payload.get("notes"),
        }
    )
    invoice_no = f"INV-OP-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    invoice_id = create_invoice(
        {
            "invoice_no": invoice_no,
            "patient_id": appointment_payload.get("patient_id"),
            "module": "OP",
            "doctor_name": appointment_payload.get("doctor_name"),
            "total_amount": amount,
            "paid_amount": 0,
            "advance_amount": 0,
            "refunded_amount": 0,
            "payment_status": "due",
            "created_by": g.current_user.get("username"),
        }
    )
    payment_mode = normalize_payment_mode(payload.get("payment_mode"))
    payment_id = record_invoice_payment(
        invoice_id,
        {
            "amount": amount,
            "payment_mode": payment_mode,
            "gateway_ref": payload.get("razorpay_payment_id"),
            "converted_from_mode": payload.get("converted_from_mode"),
            "converted_to_mode": payload.get("converted_to_mode"),
        },
    )
    log_audit_event(
        "create",
        "appointments",
        str(appointment_id),
        {
            "patient_name": appointment_payload.get("patient_name"),
            "token_no": token_no,
            "invoice_id": invoice_id,
            "invoice_payment_id": payment_id,
            "gateway_ref": payload.get("razorpay_payment_id"),
        },
    )
    return jsonify(
        {
            "appointment_id": appointment_id,
            "token_no": token_no,
            "invoice_id": invoice_id,
            "payment_id": payment_id,
        }
    )


@app.get("/api/registration/departments")
@require_permissions("patients.read")
def registration_departments_list():
    return jsonify({"departments": rows_to_dicts(list_departments(hospital_id=current_hospital_id()))})


@app.post("/api/registration/departments")
@require_permissions("patients.write")
def registration_departments_create():
    payload = request.get_json(force=True)
    try:
        department_name = normalize_department_name(payload.get("department_name"))
    except BadRequest as error:
        return jsonify({"error": str(error)}), 400

    existing = next(
        (
            row
            for row in rows_to_dicts(list_departments(hospital_id=current_hospital_id()))
            if (row.get("department_name") or "").strip().lower() == department_name.lower()
        ),
        None,
    )
    if existing:
        return jsonify({"department_id": existing.get("id"), "department_name": existing.get("department_name"), "already_exists": True})

    department_id = create_department(
        {"department_name": department_name},
        hospital_id=current_hospital_id(),
    )
    log_audit_event(
        "create",
        "departments",
        str(department_id),
        {"department_name": department_name},
    )
    return jsonify({"department_id": department_id, "department_name": department_name})


@app.get("/api/registration/consents")
@require_permissions("patients.read")
def registration_consents_list():
    patient_id = request.args.get("patient_id")
    return jsonify({"consents": rows_to_dicts(list_patient_consents(patient_id=patient_id))})


@app.post("/api/registration/consents")
@require_permissions("patients.write")
def registration_consents_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["patient_name", "consent_type", "signed_by"])
    if validation_error:
        return validation_error
    consent_id = create_patient_consent(payload)
    log_audit_event("create", "patient_consents", str(consent_id), {"patient_name": payload.get("patient_name")})
    return jsonify({"consent_id": consent_id})


@app.put("/api/registration/consents/<int:consent_id>")
@require_permissions("patients.write")
def registration_consents_update(consent_id):
    payload = request.get_json(force=True)
    updated = update_patient_consent(consent_id, payload)
    if not updated:
        return jsonify({"error": "Consent not found"}), 404
    log_audit_event("update", "patient_consents", str(consent_id), {"consent_id": consent_id})
    return jsonify({"status": "ok"})


@app.get("/api/registration/insurance")
@require_permissions("patients.read")
def registration_insurance_list():
    patient_id = request.args.get("patient_id")
    return jsonify({"verifications": rows_to_dicts(list_insurance_verifications(patient_id=patient_id))})


@app.post("/api/registration/insurance")
@require_permissions("patients.write")
def registration_insurance_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["patient_name", "insurer_name"])
    if validation_error:
        return validation_error
    verification_id = create_insurance_verification(payload)
    log_audit_event(
        "create",
        "insurance_verifications",
        str(verification_id),
        {"patient_name": payload.get("patient_name")},
    )
    return jsonify({"verification_id": verification_id})


@app.put("/api/registration/insurance/<int:verification_id>")
@require_permissions("patients.write")
def registration_insurance_update(verification_id):
    payload = request.get_json(force=True)
    updated = update_insurance_verification(verification_id, payload)
    if not updated:
        return jsonify({"error": "Insurance verification not found"}), 404
    log_audit_event(
        "update",
        "insurance_verifications",
        str(verification_id),
        {"verification_id": verification_id},
    )
    return jsonify({"status": "ok"})


@app.get("/api/op/summary")
@require_permissions("patients.read")
def op_summary():
    target_date = request.args.get("date")
    return jsonify(get_op_summary(target_date))


@app.get("/api/op/doctor-schedules")
@require_permissions("patients.read")
def op_doctor_schedules_list():
    schedule_date = request.args.get("date")
    doctor_name = request.args.get("doctor_name")
    # Optional filters: status (e.g. 'available') and department
    status = request.args.get("status") or None
    department = request.args.get("department") or None
    return jsonify(
        {
            "schedules": rows_to_dicts(
                list_doctor_schedules(
                    schedule_date=schedule_date,
                    doctor_name=doctor_name,
                    status=status,
                    department=department,
                )
            )
        }
    )


@app.post("/api/op/doctor-schedules")
@require_permissions("patients.write")
def op_doctor_schedules_create():
    payload = request.get_json(force=True)
    # Only doctor_name is required; date and time are optional (permanent roster)
    validation_error = validate_required_fields(payload, ["doctor_name"])
    if validation_error:
        return validation_error
    # Provide sensible defaults when date/time are omitted
    if not payload.get("schedule_date"):
        payload["schedule_date"] = None
    if not payload.get("start_time"):
        payload["start_time"] = "09:00"
    if not payload.get("end_time"):
        payload["end_time"] = "13:00"
    try:
        schedule_id = create_doctor_schedule(payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409
    log_audit_event(
        "create",
        "doctor_schedules",
        str(schedule_id),
        {"doctor_name": payload.get("doctor_name")},
    )
    return jsonify({"schedule_id": schedule_id})


@app.get("/api/op/doctors")
@require_permissions("patients.read")
def op_doctors_list():
    """Return all distinct doctors saved in doctor_schedules with their
    department and fee info, regardless of date. Used to populate departmentâ†’doctor
    dropdowns without any date-based filtering."""
    department = request.args.get("department") or None
    rows = list_doctor_schedules(department=department)
    seen = {}
    for row in rows:
        name = (row["doctor_name"] or "").strip()
        if not name:
            continue
        if name not in seen:
            seen[name] = {
                "doctor_name": name,
                "department": row["department"] or "",
                "consultation_fee": row["consultation_fee"] or 0,
                "review_fee": row["review_fee"] or 0,
                "status": row["status"] or "available",
            }
    return jsonify({"doctors": list(seen.values())})


@app.put("/api/op/doctor-schedules/<int:schedule_id>")
@require_permissions("patients.write")
def op_doctor_schedules_update(schedule_id):
    payload = request.get_json(force=True)
    try:
        updated = update_doctor_schedule(schedule_id, payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409
    if not updated:
        return jsonify({"error": "Doctor schedule not found"}), 404
    log_audit_event(
        "update",
        "doctor_schedules",
        str(schedule_id),
        {"schedule_id": schedule_id},
    )
    return jsonify({"status": "ok"})


@app.delete("/api/op/doctor-schedules/<int:schedule_id>")
@require_permissions("patients.write")
def op_doctor_schedules_delete(schedule_id):
    deleted = delete_doctor_schedule(schedule_id)
    if not deleted:
        return jsonify({"error": "Doctor schedule not found"}), 404
    log_audit_event(
        "delete",
        "doctor_schedules",
        str(schedule_id),
        {"schedule_id": schedule_id},
    )
    return jsonify({"status": "ok"})


@app.get("/api/patients/next-id")
@require_permissions("patients.write")
def patients_next_id():
    return jsonify(
        {"patient_id": generate_patient_id(hospital_id=current_hospital_id())}
    )



def _clean_digits(value):
    return "".join(ch for ch in str(value or "") if ch.isdigit())


_COUNTRY_PHONE_RULES = {
    "+91": ("India", 10),
    "+1": ("USA/Canada", 10),
    "+44": ("United Kingdom", 10),
    "+971": ("UAE", 9),
    "+61": ("Australia", 9),
    "+65": ("Singapore", 8),
    "+974": ("Qatar", 8),
    "+966": ("Saudi Arabia", 9),
    "+968": ("Oman", 8),
    "+965": ("Kuwait", 8),
    "+880": ("Bangladesh", 10),
    "+94": ("Sri Lanka", 9),
    "+977": ("Nepal", 10),
}

def _normalize_phone_for_country(value, country_code=None, required=False, field_label="Mobile number"):
    digits = _clean_digits(value)
    if not digits:
        if required:
            return "", f"{field_label} is required."
        return "", None

    code = str(country_code or "+91").strip() or "+91"
    country_name, exact_len = _COUNTRY_PHONE_RULES.get(code, _COUNTRY_PHONE_RULES["+91"])
    code_digits = _clean_digits(code)

    # Accept both local number (9160172353) and prefixed number (+919160172353 / 919160172353).
    if len(digits) == exact_len + len(code_digits) and digits.startswith(code_digits):
        digits = digits[len(code_digits):]

    if len(digits) != exact_len:
        return digits, f"{country_name} {code} mobile number must be exactly {exact_len} digits."
    return digits, None


def _validate_patient_registration_payload(payload):
    errors = {}
    first_name = str(payload.get("name") or "").strip()
    last_name = str(payload.get("last_name") or "").strip()
    dob = str(payload.get("dob") or "").strip()
    age = str(payload.get("age") or "").strip()
    gender = str(payload.get("gender") or "").strip()
    phone_country_code = str(payload.get("phone_country_code") or payload.get("primary_country_code") or "+91").strip() or "+91"
    family_country_code = str(payload.get("family_country_code") or "+91").strip() or "+91"
    emergency_country_code = str(payload.get("emergency_country_code") or "+91").strip() or "+91"
    phone, phone_error = _normalize_phone_for_country(payload.get("phone"), phone_country_code, required=True, field_label="Primary mobile")
    address = str(payload.get("address") or "").strip()
    family_mobile, family_mobile_error = _normalize_phone_for_country(payload.get("family_mobile"), family_country_code, required=False, field_label="Alternate mobile")
    _emergency_mobile, emergency_mobile_error = _normalize_phone_for_country(payload.get("emergency_mobile"), emergency_country_code, required=False, field_label="Emergency mobile")

    if not first_name:
        errors["name"] = "First name is required."
    if not last_name:
        errors["last_name"] = "Last name is required."
    if age:
        try:
            age_value = int(float(age))
            if age_value < 0 or age_value > 150:
                errors["age"] = "Age must be between 0 and 150."
        except (TypeError, ValueError):
            errors["age"] = "Age must be a valid number."
    if phone_error:
        errors["phone"] = phone_error
    if family_mobile_error:
        errors["family_mobile"] = family_mobile_error
    if emergency_mobile_error:
        errors["emergency_mobile"] = emergency_mobile_error

    return errors, phone, family_mobile

@app.post("/api/patients")
@require_permissions("patients.write")
def patients_create():
    payload = request.get_json(force=True)
    hospital_id = current_hospital_id()
    validation_errors, normalized_phone, normalized_family_mobile = _validate_patient_registration_payload(payload)
    if validation_errors:
        return jsonify({"error": "Patient registration validation failed.", "fields": validation_errors}), 400

    duplicate_phone = get_patient_by_phone(normalized_phone, hospital_id=hospital_id)
    if duplicate_phone:
        return (
            jsonify(
                {
                    "error": "Duplicate contact number",
                    "message": "A patient already exists with this primary contact number. Use Family Mobile Number for shared family contacts.",
                    "duplicate": row_to_dict(duplicate_phone),
                }
            ),
            409,
        )

    patient_id = generate_patient_id(hospital_id=hospital_id)
    duplicate = check_duplicate_patient(
        payload.get("name"),
        payload.get("last_name"),
        payload.get("dob"),
        normalized_phone,
        hospital_id=hospital_id,
    )
    if duplicate:
        return (
            jsonify(
                {
                    "error": "Possible duplicate",
                    "duplicate": row_to_dict(duplicate),
                }
            ),
            409,
        )

    data = {
        "patient_id": patient_id,
        "name": str(payload.get("name") or "").strip(),
        "middle_name": str(payload.get("middle_name") or "").strip() or None,
        "last_name": str(payload.get("last_name") or "").strip(),
        "dob": payload.get("dob") if payload.get("dob") else None,
        "age": int(float(payload.get("age"))) if payload.get("age") else None,
        "weight": float(payload.get("weight")) if payload.get("weight") else None,
        "height": float(payload.get("height")) if payload.get("height") else None,
        "gender": str(payload.get("gender") or "").strip() if payload.get("gender") else None,
        "pregnant": 1 if payload.get("pregnant") else 0,
        "allergies": payload.get("allergies") if payload.get("allergies") else None,
        "symptoms": payload.get("symptoms") if payload.get("symptoms") else None,
        "phone": normalized_phone,
        "address": str(payload.get("address") or "").strip() if payload.get("address") else None,
        "emergency_contact": str(payload.get("emergency_contact") or "").strip() if payload.get("emergency_contact") else None,
        "emergency_relation": str(payload.get("emergency_relation") or "").strip() if payload.get("emergency_relation") else None,
        "family_mobile": normalized_family_mobile if normalized_family_mobile else None,
    }
    add_patient(data, hospital_id=hospital_id)
    admission_id = add_admission(patient_id, "Initial registration", hospital_id=hospital_id)
    log_audit_event("create", "patients", patient_id, {"admission_id": admission_id})
    return jsonify({"patient_id": patient_id, "admission_id": admission_id})


@app.get("/api/patients/<patient_id>")
@require_permissions("patients.read")
def patients_get(patient_id):
    patient = get_patient(patient_id, hospital_id=current_hospital_id())
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    return jsonify({"patient": row_to_dict(patient)})


@app.put("/api/patients/<patient_id>")
@require_permissions("patients.write")
def patients_update(patient_id):
    hospital_id = current_hospital_id()
    payload = request.get_json(force=True)
    patient = get_patient(patient_id, hospital_id=hospital_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    data = {
        "name": payload.get("name"),
        "middle_name": payload.get("middle_name"),
        "last_name": payload.get("last_name"),
        "dob": payload.get("dob"),
        "age": payload.get("age"),
        "weight": payload.get("weight"),
        "height": payload.get("height"),
        "gender": payload.get("gender"),
        "pregnant": 1 if payload.get("pregnant") else 0,
        "allergies": payload.get("allergies"),
        "symptoms": payload.get("symptoms"),
        "phone": _clean_digits(payload.get("phone")) if payload.get("phone") is not None else patient.get("phone"),
        "address": payload.get("address", patient.get("address") if hasattr(patient, "get") else None),
        "emergency_contact": payload.get("emergency_contact", patient.get("emergency_contact") if hasattr(patient, "get") else None),
        "emergency_relation": payload.get("emergency_relation", patient.get("emergency_relation") if hasattr(patient, "get") else None),
        "family_mobile": _clean_digits(payload.get("family_mobile")) if payload.get("family_mobile") is not None else (patient.get("family_mobile") if hasattr(patient, "get") else None),
    }
    update_patient(patient_id, data)
    log_audit_event("update", "patients", patient_id, {"fields": list(data.keys())})
    return jsonify({"status": "ok"})


@app.delete("/api/patients/<patient_id>")
@require_permissions("patients.delete")
def patients_delete(patient_id):
    deleted = delete_patient(patient_id, hospital_id=current_hospital_id())
    if not deleted:
        return jsonify({"error": "Patient not found"}), 404
    log_audit_event("delete", "patients", patient_id)
    return jsonify({"status": "ok"})


@app.get("/api/patients/<patient_id>/admissions")
@require_permissions("patients.read")
def admissions_list(patient_id):
    hospital_id = current_hospital_id()
    patient = get_patient(patient_id, hospital_id=hospital_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    admissions = get_admissions(patient_id, hospital_id=hospital_id)
    return jsonify({"admissions": rows_to_dicts(admissions)})


@app.post("/api/patients/<patient_id>/admissions")
@require_permissions("patients.write")
def admissions_create(patient_id):
    hospital_id = current_hospital_id()
    patient = get_patient(patient_id, hospital_id=hospital_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    payload = request.get_json(force=True)
    notes = payload.get("notes", "")
    admission_id = add_admission(patient_id, notes)
    log_audit_event(
        "create", "admissions", str(admission_id), {"patient_id": patient_id}
    )
    return jsonify({"admission_id": admission_id})


@app.get("/api/patients/<patient_id>/documents")
@require_permissions("patients.read")
def documents_list(patient_id):
    hospital_id = current_hospital_id()
    patient = get_patient(patient_id, hospital_id=hospital_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    documents = get_documents(patient_id, hospital_id=hospital_id)
    return jsonify({"documents": rows_to_dicts(documents)})


@app.post("/api/patients/<patient_id>/documents")
@require_permissions("patients.write")
def documents_create(patient_id):
    hospital_id = current_hospital_id()
    patient = get_patient(patient_id, hospital_id=hospital_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404
    if "file" not in request.files:
        return jsonify({"error": "Missing file"}), 400
    uploaded_file = request.files["file"]
    doc_type = request.form.get("doc_type", "document")
    admission_id = request.form.get("admission_id")
    ocr_text = request.form.get("ocr_text", "")
    ocr_language = request.form.get("ocr_language", "en")
    file_bytes = uploaded_file.read()
    uploaded_file.stream.seek(0)

    filepath = save_uploaded_file(uploaded_file, file_bytes, doc_type)
    document_id = add_document(
        patient_id,
        admission_id if admission_id else None,
        doc_type,
        filepath,
        ocr_text,
        ocr_language,
        file_name=uploaded_file.filename,
        mime_type=uploaded_file.mimetype,
        file_data=None,
        hospital_id=hospital_id,
    )
    return jsonify(
        {"document_id": document_id, "file_path": filepath, "stored_in_db": False}
    )


@app.get("/api/documents/<int:document_id>/file")
@require_permissions("patients.read")
def document_file(document_id):
    document = get_document(document_id, hospital_id=current_hospital_id())
    if not document:
        return jsonify({"error": "Document not found"}), 404

    mime_type = document["mime_type"] or "application/octet-stream"
    file_name = (
        document["file_name"]
        or os.path.basename(document["file_path"] or "")
        or "document.bin"
    )

    file_data = document["file_data"]
    if file_data:
        return send_file(
            io.BytesIO(file_data),
            mimetype=mime_type,
            as_attachment=False,
            download_name=file_name,
        )

    file_path = document["file_path"]
    file_bytes = STORAGE.read(file_path)
    if file_bytes:
        return send_file(
            io.BytesIO(file_bytes),
            mimetype=mime_type,
            as_attachment=False,
            download_name=file_name,
        )

    return jsonify({"error": "Document file content unavailable"}), 404


@app.delete("/api/documents/<int:document_id>")
@require_permissions("patients.write")
def document_delete(document_id):
    hospital_id = current_hospital_id()
    document = get_document(document_id, hospital_id=hospital_id)
    if not document:
        return jsonify({"error": "Document not found"}), 404

    deleted = delete_document(document_id, hospital_id=hospital_id)
    if not deleted:
        return jsonify({"error": "Document not found"}), 404

    STORAGE.delete(document["file_path"])

    return jsonify({"status": "ok"})


@app.post("/api/documents/<int:document_id>/ocr")
@require_permissions("patients.write")
def document_process_ocr(document_id):
    hospital_id = current_hospital_id()
    document = get_document(document_id, hospital_id=hospital_id)
    if not document:
        return jsonify({"error": "Document not found"}), 404

    payload = request.get_json(silent=True)
    payload = payload if isinstance(payload, dict) else {}
    requested_language = payload.get("language") or document["ocr_language"] or "en"
    file_name = (
        document["file_name"]
        or os.path.basename(document["file_path"] or "")
        or "document"
    )

    file_bytes = document["file_data"]
    if not file_bytes:
        file_path = document["file_path"]
        if not file_path:
            return jsonify({"error": "Document file content unavailable"}), 404
        file_bytes = STORAGE.read(file_path)
        if not file_bytes:
            return jsonify({"error": "Document file content unavailable"}), 404

    ocr_text = extract_text_from_image(
        file_bytes, requested_language, document["doc_type"], filename=file_name
    )
    if isinstance(ocr_text, str) and ocr_text.startswith("OCR Error:"):
        return jsonify({"error": ocr_text}), 400

    updated = update_document_ocr(
        document_id, ocr_text, requested_language, hospital_id=hospital_id
    )
    if not updated:
        return jsonify({"error": "Document not found"}), 404

    return jsonify(
        {
            "document_id": document_id,
            "ocr_text": ocr_text,
            "ocr_language": requested_language,
            "updated": True,
        }
    )


@app.post("/api/ocr")
@require_permissions("patients.write")
def ocr_extract():
    if "file" not in request.files:
        return jsonify({"error": "Missing file"}), 400
    uploaded_file = request.files["file"]
    language = request.form.get("language", "en")
    doc_type = request.form.get("doc_type", "document")
    content = uploaded_file.read()
    text = extract_text_from_image(
        content, language, doc_type, filename=uploaded_file.filename
    )
    return jsonify({"text": text})


@app.post("/api/export/pdf")
@require_permissions("patients.write")
def export_pdf():
    payload = request.get_json(force=True)
    pdf_bytes = generate_pdf(
        payload.get("patient_name", ""),
        payload.get("doc_type", "document"),
        payload.get("ocr_text", ""),
        payload.get("date"),
    )
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name="document.pdf",
    )


@app.post("/api/export/word")
@require_permissions("patients.write")
def export_word():
    payload = request.get_json(force=True)
    word_bytes = generate_word(
        payload.get("patient_name", ""),
        payload.get("doc_type", "document"),
        payload.get("ocr_text", ""),
        payload.get("date"),
    )
    return send_file(
        io.BytesIO(word_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        as_attachment=True,
        download_name="document.docx",
    )


@app.get("/api/export/patients/csv")
@require_permissions("patients.read")
def export_patients_csv():
    query = (request.args.get("q") or "").strip()
    hospital_id = current_hospital_id()
    patients = (
        search_patients(query, hospital_id=hospital_id)
        if query
        else get_all_patients(hospital_id=hospital_id)
    )

    csv_stream = io.StringIO()
    writer = csv.writer(csv_stream)
    headers = [
        "patient_id",
        "name",
        "middle_name",
        "last_name",
        "dob",
        "age",
        "weight",
        "height",
        "gender",
        "pregnant",
        "allergies",
        "symptoms",
        "phone",
        "created_at",
        "updated_at",
    ]
    writer.writerow(headers)
    for row in patients:
        item = row_to_dict(row)
        writer.writerow(
            [
                item.get(header, "") if item.get(header) is not None else ""
                for header in headers
            ]
        )

    filename_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        io.BytesIO(csv_stream.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"patients_{filename_suffix}.csv",
    )


# ==================== Employee Management ====================


@app.get("/api/employees")
@require_permissions("admin.use")
def employees_list():
    employees = get_all_employees(hospital_id=current_hospital_id())
    return jsonify({"employees": rows_to_dicts(employees)})


@app.post("/api/employees")
@require_permissions("admin.use")
def employees_create():
    payload = request.get_json(force=True)
    result = signup_employee(payload, allow_admin_creation=True, hospital_id=current_hospital_id())
    status = 201 if result.get("success") else 400
    return jsonify(result), status


@app.get("/api/employees/stats")
@require_permissions("admin.use")
def employees_stats():
    return jsonify(get_employee_stats(hospital_id=current_hospital_id()))


@app.get("/api/employees/search")
@require_permissions("admin.use")
def employees_search():
    query = request.args.get("q")
    if not query:
        return jsonify({"employees": []})
    employees = search_employees(query, hospital_id=current_hospital_id())
    return jsonify({"employees": rows_to_dicts(employees)})


@app.route("/api/employees/<employee_id>", methods=["GET", "PUT", "DELETE"])
@require_permissions("admin.use")
def employees_detail(employee_id):
    hospital_id = current_hospital_id()
    if request.method == "GET":
        employee = get_employee(employee_id=employee_id, hospital_id=hospital_id)
        if not employee:
            return jsonify({"error": "Employee not found"}), 404
        return jsonify({"employee": row_to_dict(employee)})

    if request.method == "PUT":
        user_permissions = resolve_user_permissions(g.current_user)
        if "employees.write" not in user_permissions:
            return jsonify(
                {"error": "Forbidden", "required_permissions": ["employees.write"]}
            ), 403
        payload = request.get_json(force=True)
        updated = update_employee(employee_id, payload, hospital_id=hospital_id)
        if not updated:
            return jsonify({"error": "Employee not found"}), 404
        return jsonify({"status": "ok"})

    # DELETE
    user_permissions = resolve_user_permissions(g.current_user)
    if "employees.write" not in user_permissions:
        return jsonify(
            {"error": "Forbidden", "required_permissions": ["employees.write"]}
        ), 403
    deleted = delete_employee(employee_id)
    if not deleted:
        return jsonify({"error": "Employee not found"}), 404
    return jsonify({"status": "ok"})


@app.post("/api/employees/<employee_id>/activate")
@require_permissions("admin.use")
def employees_activate(employee_id):
    activated = activate_employee(employee_id, hospital_id=current_hospital_id())
    if not activated:
        return jsonify({"error": "Employee not found"}), 404
    return jsonify({"status": "active"})


@app.post("/api/employees/<employee_id>/deactivate")
@require_permissions("admin.use")
def employees_deactivate(employee_id):
    deactivated = deactivate_employee(employee_id, hospital_id=current_hospital_id())
    if not deactivated:
        return jsonify({"error": "Employee not found"}), 404
    return jsonify({"status": "inactive"})


# ==================== Extended Patient Management ====================


@app.get("/api/patients/<patient_id>/encounters")
@require_permissions("patients.read")
def patient_encounters(patient_id):
    return jsonify(
        {"encounters": rows_to_dicts(list_encounters(patient_id=patient_id))}
    )


@app.post("/api/patients/<patient_id>/encounters")
@require_permissions("patients.write")
def patient_encounter_create(patient_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["encounter_type"])
    if validation_error:
        return validation_error
    encounter_id = create_encounter(
        {
            "patient_id": patient_id,
            "encounter_type": payload.get("encounter_type", "OP"),
            "insurance_provider": payload.get("insurance_provider"),
            "insurance_policy_no": payload.get("insurance_policy_no"),
            "is_accident": payload.get("is_accident", False),
            "referral_source": payload.get("referral_source"),
            "referral_name": payload.get("referral_name"),
            "status": payload.get("status", "active"),
            "created_by": g.current_user.get("username"),
        }
    )
    log_audit_event(
        "create", "encounters", str(encounter_id), {"patient_id": patient_id}
    )
    return jsonify({"encounter_id": encounter_id})


@app.get("/api/patients/<patient_id>/beds")
@require_permissions("patients.read")
def patient_beds(patient_id):
    active_only = request.args.get("active_only", "false").lower() == "true"
    return jsonify(
        {
            "beds": rows_to_dicts(
                list_bed_allocations(patient_id=patient_id, active_only=active_only)
            )
        }
    )


@app.post("/api/patients/<patient_id>/beds")
@require_permissions("patients.write")
def patient_bed_assign(patient_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload, ["admission_id", "ward", "room_no", "bed_no"]
    )
    if validation_error:
        return validation_error
    bed_id = assign_bed(
        {
            "admission_id": payload.get("admission_id"),
            "patient_id": patient_id,
            "ward": payload.get("ward"),
            "room_no": payload.get("room_no"),
            "bed_no": payload.get("bed_no"),
            "status": payload.get("status", "active"),
        }
    )
    log_audit_event(
        "create", "bed_allocations", str(bed_id), {"patient_id": patient_id}
    )
    return jsonify({"bed_allocation_id": bed_id})


def get_or_create_active_admission(patient_id, hospital_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM admissions WHERE patient_id = ? AND hospital_id = ? AND discharge_date IS NULL ORDER BY admission_date DESC LIMIT 1",
            (patient_id, hospital_id)
        )
        row = cursor.fetchone()
        if row:
            return row[0]
        
        cursor.execute(
            "SELECT id FROM admissions WHERE patient_id = ? AND hospital_id = ? ORDER BY admission_date DESC LIMIT 1",
            (patient_id, hospital_id)
        )
        row = cursor.fetchone()
        if row:
            return row[0]
            
        return add_admission(patient_id, "Bed allocation admission", hospital_id=hospital_id)


@app.get("/api/bed/config")
@require_permissions("patients.read")
def get_bed_config():
    hospital_id = current_hospital_id()
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT capacity FROM bed_config WHERE hospital_id = ?", (hospital_id,))
        row = cursor.fetchone()
        capacity = row[0] if row else 50
        
        cursor.execute(
            "SELECT COUNT(*) FROM bed_allocations WHERE status = 'active' AND admission_id IN (SELECT id FROM admissions WHERE hospital_id = ?)",
            (hospital_id,)
        )
        occupied_count = cursor.fetchone()[0]
        
    return jsonify({"capacity": capacity, "occupied_count": occupied_count})


@app.post("/api/bed/config")
@require_permissions("patients.write")
def update_bed_config():
    payload = request.get_json(force=True)
    capacity = int(payload.get("capacity", 50))
    hospital_id = current_hospital_id()
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT capacity FROM bed_config WHERE hospital_id = ?", (hospital_id,))
        if cursor.fetchone():
            cursor.execute("UPDATE bed_config SET capacity = ? WHERE hospital_id = ?", (capacity, hospital_id))
        else:
            cursor.execute("INSERT INTO bed_config (hospital_id, capacity) VALUES (?, ?)", (hospital_id, capacity))
        conn.commit()
    return jsonify({"status": "success", "capacity": capacity})


@app.get("/api/bed/occupancy")
@require_permissions("patients.read")
def get_bed_occupancy():
    hospital_id = current_hospital_id()
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT b.id, b.patient_id, b.ward, b.room_no, b.bed_no, b.amount_per_day, b.total_days, b.total_amount, b.allocated_at,
                   p.name, p.last_name
            FROM bed_allocations b
            JOIN patients p ON b.patient_id = p.patient_id
            WHERE b.status = 'active' AND b.admission_id IN (SELECT id FROM admissions WHERE hospital_id = ?)
            ORDER BY b.allocated_at DESC
            """,
            (hospital_id,)
        )
        rows = cursor.fetchall()
        occupied = []
        for r in rows:
            occupied.append({
                "id": r[0],
                "patient_id": r[1],
                "ward": r[2],
                "room_no": r[3],
                "bed_no": r[4],
                "amount_per_day": r[5],
                "total_days": r[6],
                "total_amount": r[7],
                "allocated_at": r[8],
                "patient_name": f"{r[9]} {r[10]}".strip()
            })
    return jsonify({"occupied": occupied})


@app.post("/api/bed/allot")
@require_permissions("patients.write")
def allot_bed_api():
    payload = request.get_json(force=True)
    patient_id = payload.get("patient_id")
    ward = payload.get("ward")
    room_no = payload.get("room_no")
    bed_no = payload.get("bed_no")
    amount_per_day = float(payload.get("amount_per_day") or 0)
    total_days = int(payload.get("total_days") or 0)
    total_amount = float(payload.get("total_amount") or 0)
    
    if not patient_id or not ward or not room_no or not bed_no:
        return jsonify({"error": "Missing required fields"}), 400
        
    hospital_id = current_hospital_id()
    
    with get_connection() as conn:
        cursor = conn.cursor()
        if len(patient_id) == 4 and patient_id.isdigit():
            cursor.execute(
                "SELECT patient_id FROM patients WHERE hospital_id = ? AND patient_id LIKE ?",
                (hospital_id, f"%{patient_id}")
            )
            row = cursor.fetchone()
            if row:
                patient_id = row[0]
            else:
                return jsonify({"error": "Patient not found matching those last 4 digits"}), 404
        else:
            cursor.execute(
                "SELECT patient_id FROM patients WHERE hospital_id = ? AND patient_id = ?",
                (hospital_id, patient_id)
            )
            if not cursor.fetchone():
                return jsonify({"error": "Patient not found"}), 404
                
    admission_id = get_or_create_active_admission(patient_id, hospital_id)
    
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE bed_allocations SET status = 'released', released_at = CURRENT_TIMESTAMP WHERE patient_id = ? AND status = 'active'",
            (patient_id,)
        )
        cursor.execute(
            """
            INSERT INTO bed_allocations (
                admission_id, patient_id, ward, room_no, bed_no, amount_per_day, total_days, total_amount, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active')
            """,
            (admission_id, patient_id, ward, room_no, bed_no, amount_per_day, total_days, total_amount)
        )
        bed_id = cursor.lastrowid
        conn.commit()
        
    log_audit_event("create", "bed_allocations", str(bed_id), {"patient_id": patient_id})
    return jsonify({"status": "success", "bed_allocation_id": bed_id})


@app.post("/api/bed/release/<int:allocation_id>")
@require_permissions("patients.write")
def release_bed_api(allocation_id):
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE bed_allocations SET status = 'released', released_at = CURRENT_TIMESTAMP WHERE id = ?",
            (allocation_id,)
        )
        conn.commit()
    return jsonify({"status": "success"})


@app.get("/api/bed/history")
@require_permissions("patients.read")
def get_bed_history_all():
    patient_id = request.args.get("patient_id")
    with get_connection() as conn:
        cursor = conn.cursor()
        query = """
            SELECT b.id, b.patient_id, b.ward, b.room_no, b.bed_no, b.amount_per_day, b.total_days, b.total_amount, b.allocated_at, b.released_at, b.status,
                   p.name, p.last_name
            FROM bed_allocations b
            JOIN patients p ON b.patient_id = p.patient_id
        """
        params = []
        if patient_id:
            query += " WHERE b.patient_id = ?"
            params.append(patient_id)
        query += " ORDER BY b.allocated_at DESC"
        
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        history = []
        for r in rows:
            history.append({
                "id": r[0],
                "patient_id": r[1],
                "ward": r[2],
                "room_no": r[3],
                "bed_no": r[4],
                "amount_per_day": r[5],
                "total_days": r[6],
                "total_amount": r[7],
                "allocated_at": r[8],
                "released_at": r[9],
                "status": r[10],
                "patient_name": f"{r[11]} {r[12]}".strip()
            })
    return jsonify({"history": history})


@app.get("/api/patients/<patient_id>/medications")
@require_permissions("patients.read")
def patient_medications(patient_id):
    pending_only = request.args.get("pending_only", "false").lower() == "true"
    return jsonify(
        {
            "medications": rows_to_dicts(
                list_medication_schedules(patient_id, pending_only=pending_only)
            )
        }
    )


@app.post("/api/patients/<patient_id>/medications")
@require_permissions("patients.write")
def patient_medications_create(patient_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["medicine_name"])
    if validation_error:
        return validation_error
    schedule_id = add_medication_schedule(
        {
            "patient_id": patient_id,
            "medicine_name": payload.get("medicine_name"),
            "dosage": payload.get("dosage"),
            "schedule_time": payload.get("schedule_time")
            or current_ist_datetime().isoformat(timespec="seconds"),
            "administered": payload.get("administered", False),
            "alert_enabled": payload.get("alert_enabled", True),
            "notes": payload.get("notes"),
        }
    )
    log_audit_event(
        "create", "medication_schedules", str(schedule_id), {"patient_id": patient_id}
    )
    return jsonify({"schedule_id": schedule_id})


@app.get("/api/patients/<patient_id>/notes")
@require_permissions("patients.read")
def patient_notes(patient_id):
    return jsonify({"notes": rows_to_dicts(list_observation_notes(patient_id))})


@app.post("/api/patients/<patient_id>/notes")
@require_permissions("patients.write")
def patient_notes_create(patient_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["note"])
    if validation_error:
        return validation_error
    note_id = add_observation_note(
        {
            "patient_id": patient_id,
            "admission_id": payload.get("admission_id"),
            "doctor_name": payload.get("doctor_name"),
            "note": payload.get("note"),
            "treatment_plan": payload.get("treatment_plan"),
        }
    )
    log_audit_event(
        "create", "observation_notes", str(note_id), {"patient_id": patient_id}
    )
    return jsonify({"note_id": note_id})


@app.get("/api/patients/<patient_id>/certificates")
@require_permissions("patients.read")
def patient_certificates(patient_id):
    return jsonify({"certificates": rows_to_dicts(list_certificates(patient_id))})


@app.post("/api/patients/<patient_id>/certificates")
@require_permissions("patients.write")
def patient_certificates_create(patient_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload, ["certificate_type", "title", "body"]
    )
    if validation_error:
        return validation_error
    certificate_id = create_certificate(
        {
            "patient_id": patient_id,
            "admission_id": payload.get("admission_id"),
            "certificate_type": payload.get("certificate_type"),
            "title": payload.get("title"),
            "body": payload.get("body"),
            "issued_by": g.current_user.get("username"),
        }
    )
    log_audit_event(
        "create",
        "certificates",
        str(certificate_id),
        {"patient_id": patient_id, "certificate_type": payload.get("certificate_type")},
    )
    return jsonify({"certificate_id": certificate_id})


@app.delete("/api/certificates/<int:certificate_id>")
@require_permissions("patients.write")
def patient_certificates_delete(certificate_id):
    deleted = delete_certificate(certificate_id)
    if not deleted:
        return jsonify({"error": "Certificate not found"}), 404
    log_audit_event("delete", "certificates", str(certificate_id), {"certificate_id": certificate_id})
    return jsonify({"status": "ok"})


@app.get("/api/patients/<patient_id>/movements")
@require_permissions("patients.read")
def patient_movements(patient_id):
    return jsonify({"movements": rows_to_dicts(list_patient_movements(patient_id))})


@app.post("/api/patients/<patient_id>/movements")
@require_permissions("patients.write")
def patient_movements_create(patient_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["to_department"])
    if validation_error:
        return validation_error
    movement_id = add_patient_movement(
        {
            "patient_id": patient_id,
            "admission_id": payload.get("admission_id"),
            "from_department": payload.get("from_department"),
            "to_department": payload.get("to_department"),
            "moved_by": g.current_user.get("username"),
        }
    )
    log_audit_event(
        "create", "patient_movements", str(movement_id), {"patient_id": patient_id}
    )
    return jsonify({"movement_id": movement_id})


# ==================== Billing ====================


@app.get("/api/billing/invoices")
@require_permissions("billing.read")
def billing_invoices():
    patient_id = request.args.get("patient_id")
    module = request.args.get("module")
    return jsonify(
        {"invoices": rows_to_dicts(list_invoices(patient_id=patient_id, module=module))}
    )


@app.post("/api/billing/invoices")
@require_permissions("billing.write")
def billing_create_invoice():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["module", "total_amount"])
    if validation_error:
        return validation_error
    invoice_no = (
        payload.get("invoice_no") or f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    )
    invoice_id = create_invoice(
        {
            "invoice_no": invoice_no,
            "patient_id": payload.get("patient_id"),
            "module": payload.get("module", "OP"),
            "doctor_name": payload.get("doctor_name"),
            "clinic_name": payload.get("clinic_name"),
            "referral_source": payload.get("referral_source"),
            "subtotal": payload.get("subtotal", 0),
            "tax": payload.get("tax", 0),
            "discount": payload.get("discount", 0),
            "total_amount": payload.get("total_amount", 0),
            "paid_amount": payload.get("paid_amount", 0),
            "advance_amount": payload.get("advance_amount", 0),
            "refunded_amount": payload.get("refunded_amount", 0),
            "payment_status": payload.get("payment_status", "due"),
            "created_by": g.current_user.get("username"),
        }
    )
    log_audit_event(
        "create", "billing_invoices", str(invoice_id), {"invoice_no": invoice_no}
    )
    return jsonify({"invoice_id": invoice_id, "invoice_no": invoice_no})


@app.put("/api/billing/invoices/<int:invoice_id>")
@require_permissions("billing.write")
def billing_update_invoice(invoice_id):
    payload = request.get_json(force=True)
    updated = update_invoice(invoice_id, payload)
    if not updated:
        return jsonify({"error": "Invoice not found"}), 404
    log_audit_event(
        "update", "billing_invoices", str(invoice_id), {"invoice_id": invoice_id}
    )
    return jsonify({"status": "ok"})


@app.delete("/api/billing/invoices/<int:invoice_id>")
@require_permissions("billing.write")
def billing_delete_invoice(invoice_id):
    deleted = delete_invoice(invoice_id)
    if not deleted:
        return jsonify({"error": "Invoice not found"}), 404
    log_audit_event(
        "delete", "billing_invoices", str(invoice_id), {"invoice_id": invoice_id}
    )
    return jsonify({"status": "ok"})


@app.post("/api/billing/razorpay/order")
@require_permissions("billing.write")
def billing_razorpay_order():
    config_error = require_razorpay_configured()
    if config_error:
        return config_error
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["invoice_id", "amount"])
    if validation_error:
        return validation_error

    amount_paise = to_amount_paise(payload.get("amount"))
    if amount_paise is None:
        return jsonify({"error": "Amount must be greater than zero."}), 400

    try:
        invoice_id = int(payload.get("invoice_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "invoice_id must be a valid number."}), 400
    if not get_invoice_by_id(invoice_id):
        return jsonify({"error": "Invoice not found"}), 404
    receipt = payload.get("receipt") or f"bill-{invoice_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    notes = payload.get("notes") if isinstance(payload.get("notes"), dict) else {}
    notes = {
        **notes,
        "invoice_id": str(invoice_id),
        "hospital_id": str(current_hospital_id()),
    }
    try:
        order = create_razorpay_order(
            amount_paise=amount_paise,
            currency=payload.get("currency", "INR"),
            receipt=receipt,
            notes=notes,
        )
    except RuntimeError as error:
        return jsonify({"error": str(error)}), 502

    return jsonify(
        {
            "key_id": RAZORPAY_KEY_ID,
            "order_id": order.get("id"),
            "amount": order.get("amount"),
            "currency": order.get("currency"),
            "receipt": order.get("receipt"),
        }
    )


@app.post("/api/billing/razorpay/verify")
@require_permissions("billing.write")
def billing_razorpay_verify():
    config_error = require_razorpay_configured()
    if config_error:
        return config_error
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload,
        [
            "invoice_id",
            "amount",
            "razorpay_order_id",
            "razorpay_payment_id",
            "razorpay_signature",
        ],
    )
    if validation_error:
        return validation_error

    if not verify_razorpay_signature(
        payload.get("razorpay_order_id"),
        payload.get("razorpay_payment_id"),
        payload.get("razorpay_signature"),
    ):
        return jsonify({"error": "Invalid Razorpay signature."}), 400

    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than zero."}), 400

    try:
        invoice_id = int(payload.get("invoice_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "invoice_id must be a valid number."}), 400

    payment_id = record_invoice_payment(
        invoice_id,
        {
            "amount": amount,
            "payment_mode": normalize_payment_mode(payload.get("payment_mode")),
            "gateway_ref": payload.get("razorpay_payment_id"),
            "converted_from_mode": payload.get("converted_from_mode"),
            "converted_to_mode": payload.get("converted_to_mode"),
        },
    )
    if payment_id is None:
        return jsonify({"error": "Invoice not found"}), 404
    log_audit_event(
        "create",
        "billing_payments",
        str(payment_id),
        {
            "invoice_id": invoice_id,
            "gateway_ref": payload.get("razorpay_payment_id"),
            "razorpay_order_id": payload.get("razorpay_order_id"),
        },
    )
    return jsonify({"payment_id": payment_id})


@app.post("/api/billing/invoices/<int:invoice_id>/payments")
@require_permissions("billing.write")
def billing_record_payment(invoice_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["amount", "payment_mode"])
    if validation_error:
        return validation_error
    payment_id = record_invoice_payment(
        invoice_id,
        {
            "amount": payload.get("amount", 0),
            "payment_mode": payload.get("payment_mode", "cash"),
            "gateway_ref": payload.get("gateway_ref"),
            "converted_from_mode": payload.get("converted_from_mode"),
            "converted_to_mode": payload.get("converted_to_mode"),
        },
    )
    if payment_id is None:
        return jsonify({"error": "Invoice not found"}), 404
    log_audit_event(
        "create", "billing_payments", str(payment_id), {"invoice_id": invoice_id}
    )
    return jsonify({"payment_id": payment_id})


@app.post("/api/billing/direct-payment")
@require_permissions("billing.write")
def billing_direct_payment():
    payload = request.get_json(force=True) or {}
    validation_error = validate_required_fields(payload, ["patient_id", "amount", "payment_mode"])
    if validation_error:
        return validation_error
    amount = float(payload.get("amount") or 0)
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than zero."}), 400

    payment_for = str(payload.get("payment_for") or payload.get("module") or "OP").strip()
    module = normalize_billing_module(payment_for)
    invoice_no = f"PAY-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    invoice_id = create_invoice(
        {
            "invoice_no": invoice_no,
            "patient_id": payload.get("patient_id"),
            "module": module,
            "clinic_name": payload.get("patient_name"),
            "referral_source": payment_for,
            "total_amount": amount,
            "paid_amount": 0,
            "advance_amount": 0,
            "refunded_amount": 0,
            "payment_status": "due",
            "created_by": g.current_user.get("username"),
        }
    )
    payment_id = record_invoice_payment(
        invoice_id,
        {
            "amount": amount,
            "payment_mode": normalize_payment_mode(payload.get("payment_mode")),
            "gateway_ref": payload.get("gateway_ref"),
            "converted_from_mode": payload.get("converted_from_mode"),
            "converted_to_mode": payload.get("converted_to_mode"),
        },
    )
    log_audit_event(
        "create",
        "billing_direct_payment",
        str(payment_id),
        {"invoice_id": invoice_id, "patient_id": payload.get("patient_id"), "payment_for": payment_for},
    )
    return jsonify({"invoice_id": invoice_id, "invoice_no": invoice_no, "payment_id": payment_id, "module": module})


@app.get("/api/billing/revenue-summary")
@require_permissions("billing.read")
def billing_revenue_summary():
    collection_date = request.args.get("date")
    collection_month = request.args.get("month")
    return jsonify(get_revenue_summary(collection_date=collection_date, collection_month=collection_month))


@app.get("/api/billing/claims")
@require_permissions("billing.read")
def billing_claims():
    invoice_id = request.args.get("invoice_id", type=int)
    status = request.args.get("status")
    return jsonify(
        {"claims": rows_to_dicts(list_insurance_claims(invoice_id=invoice_id, status=status))}
    )


@app.post("/api/billing/claims")
@require_permissions("billing.write")
def billing_create_claim():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["invoice_id", "insurer_name", "claim_amount"])
    if validation_error:
        return validation_error
    claim_id = create_insurance_claim(payload)
    log_audit_event("create", "insurance_claims", str(claim_id), {"invoice_id": payload.get("invoice_id")})
    return jsonify({"claim_id": claim_id})


@app.put("/api/billing/claims/<int:claim_id>")
@require_permissions("billing.write")
def billing_update_claim(claim_id):
    payload = request.get_json(force=True)
    updated = update_insurance_claim(claim_id, payload)
    if not updated:
        return jsonify({"error": "Claim not found"}), 404
    log_audit_event("update", "insurance_claims", str(claim_id), {"claim_id": claim_id})
    return jsonify({"status": "ok"})


@app.delete("/api/billing/claims/<int:claim_id>")
@require_permissions("billing.write")
def billing_delete_claim(claim_id):
    deleted = delete_insurance_claim(claim_id)
    if not deleted:
        return jsonify({"error": "Claim not found"}), 404
    log_audit_event("delete", "insurance_claims", str(claim_id), {"claim_id": claim_id})
    return jsonify({"status": "ok"})


# ==================== Pharmacy (module removed) ====================


@app.route("/api/pharmacy", defaults={"_path": ""}, methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])
@app.route("/api/pharmacy/<path:_path>", methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])
def pharmacy_module_removed(_path):
    if request.method == "OPTIONS":
        return ("", 204)
    return (
        jsonify(
            {
                "error": "Pharmacy module removed",
                "message": "The Pharmacy module has been removed from HospAI. Existing pharmacy records are preserved but pharmacy operations are unavailable.",
            }
        ),
        410,
    )


# ==================== Lab & Diagnostics ====================


@app.get("/api/lab/vendors")
@require_permissions("lab.read")
def lab_vendors_list():
    return jsonify({"vendors": rows_to_dicts(list_lab_vendors())})


@app.post("/api/lab/vendors")
@require_permissions("lab.write")
def lab_vendors_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["vendor_name"])
    if validation_error:
        return validation_error
    vendor_id = create_lab_vendor(payload)
    log_audit_event(
        "create",
        "lab_vendors",
        str(vendor_id),
        {"vendor_name": payload.get("vendor_name")},
    )
    return jsonify({"vendor_id": vendor_id})


@app.put("/api/lab/vendors/<int:vendor_id>")
@require_permissions("lab.write")
def lab_vendors_update(vendor_id):
    payload = request.get_json(force=True)
    updated = update_lab_vendor(vendor_id, payload)
    if not updated:
        return jsonify({"error": "Vendor not found"}), 404
    log_audit_event("update", "lab_vendors", str(vendor_id), {"vendor_id": vendor_id})
    return jsonify({"status": "ok"})


@app.delete("/api/lab/vendors/<int:vendor_id>")
@require_permissions("lab.write")
def lab_vendors_delete(vendor_id):
    deleted = delete_lab_vendor(vendor_id)
    if not deleted:
        return jsonify({"error": "Vendor not found"}), 404
    log_audit_event("delete", "lab_vendors", str(vendor_id), {"vendor_id": vendor_id})
    return jsonify({"status": "ok"})


@app.get("/api/lab/diagnostics")
@require_permissions("lab.read")
def lab_diagnostics_list():
    patient_id = request.args.get("patient_id")
    doctor_name = request.args.get("doctor_name")
    if patient_id:
        patient = get_patient(patient_id, hospital_id=current_hospital_id())
        patient_id = patient["patient_id"] if patient else patient_id
    return jsonify(
        {
            "diagnostics": rows_to_dicts(
                list_diagnostics(patient_id=patient_id, doctor_name=doctor_name)
            )
        }
    )


@app.post("/api/lab/diagnostics")
@require_permissions("lab.write")
def lab_diagnostics_create():
    payload = request.get_json(force=True)
    if payload.get("patient_id"):
        patient = get_patient(payload.get("patient_id"), hospital_id=current_hospital_id())
        if patient:
            payload["patient_id"] = patient["patient_id"]
    validation_error = validate_required_fields(payload, ["test_name", "amount"])
    if validation_error:
        return validation_error
    diagnostic_id = create_diagnostic_record(payload)
    log_audit_event(
        "create",
        "diagnostics",
        str(diagnostic_id),
        {"test_name": payload.get("test_name")},
    )
    return jsonify({"diagnostic_id": diagnostic_id})


@app.put("/api/lab/diagnostics/<int:diagnostic_id>")
@require_permissions("lab.write")
def lab_diagnostics_update(diagnostic_id):
    payload = request.get_json(force=True)
    updated = update_diagnostic_record(diagnostic_id, payload)
    if not updated:
        return jsonify({"error": "Diagnostic record not found"}), 404
    log_audit_event(
        "update", "diagnostics", str(diagnostic_id), {"diagnostic_id": diagnostic_id}
    )
    return jsonify({"status": "ok"})


@app.delete("/api/lab/diagnostics/<int:diagnostic_id>")
@require_permissions("lab.write")
def lab_diagnostics_delete(diagnostic_id):
    deleted = delete_diagnostic_record(diagnostic_id)
    if not deleted:
        return jsonify({"error": "Diagnostic record not found"}), 404
    log_audit_event(
        "delete", "diagnostics", str(diagnostic_id), {"diagnostic_id": diagnostic_id}
    )
    return jsonify({"status": "ok"})


@app.get("/api/lab/summary")
@require_permissions("lab.read")
def lab_summary():
    return jsonify(get_diagnostic_summary())


# ==================== OT ====================


@app.get("/api/ot/theatres")
@require_permissions("ot.read")
def ot_theatres_list():
    return jsonify({"theatres": rows_to_dicts(list_ot_theatres())})


@app.post("/api/ot/theatres")
@require_permissions("ot.write")
def ot_theatres_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["theatre_code", "theatre_name"])
    if validation_error:
        return validation_error
    theatre_id = create_ot_theatre(payload)
    log_audit_event("create", "ot_theatres", str(theatre_id), {"theatre_code": payload.get("theatre_code")})
    return jsonify({"theatre_id": theatre_id})


@app.put("/api/ot/theatres/<int:theatre_id>")
@require_permissions("ot.write")
def ot_theatres_update(theatre_id):
    payload = request.get_json(force=True)
    updated = update_ot_theatre(theatre_id, payload)
    if not updated:
        return jsonify({"error": "Theatre not found"}), 404
    log_audit_event("update", "ot_theatres", str(theatre_id), {"theatre_id": theatre_id})
    return jsonify({"status": "ok"})


@app.delete("/api/ot/theatres/<int:theatre_id>")
@require_permissions("ot.write")
def ot_theatres_delete(theatre_id):
    deleted = delete_ot_theatre(theatre_id)
    if not deleted:
        return jsonify({"error": "Theatre not found"}), 404
    log_audit_event("delete", "ot_theatres", str(theatre_id), {"theatre_id": theatre_id})
    return jsonify({"status": "ok"})


@app.get("/api/ot/surgeries")
@require_permissions("ot.read")
def ot_surgeries_list():
    theatre_id = request.args.get("theatre_id", type=int)
    status = request.args.get("status")
    return jsonify({"surgeries": rows_to_dicts(list_ot_surgeries(theatre_id=theatre_id, status=status))})


@app.post("/api/ot/surgeries")
@require_permissions("ot.write")
def ot_surgeries_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["theatre_id", "procedure_name", "surgeon_name", "scheduled_start"])
    if validation_error:
        return validation_error
    surgery_id = create_ot_surgery(payload)
    log_audit_event("create", "ot_surgeries", str(surgery_id), {"procedure_name": payload.get("procedure_name")})
    return jsonify({"surgery_id": surgery_id})


@app.put("/api/ot/surgeries/<int:surgery_id>")
@require_permissions("ot.write")
def ot_surgeries_update(surgery_id):
    payload = request.get_json(force=True)
    updated = update_ot_surgery(surgery_id, payload)
    if not updated:
        return jsonify({"error": "Surgery not found"}), 404
    log_audit_event("update", "ot_surgeries", str(surgery_id), {"surgery_id": surgery_id})
    return jsonify({"status": "ok"})


@app.delete("/api/ot/surgeries/<int:surgery_id>")
@require_permissions("ot.write")
def ot_surgeries_delete(surgery_id):
    deleted = delete_ot_surgery(surgery_id)
    if not deleted:
        return jsonify({"error": "Surgery not found"}), 404
    log_audit_event("delete", "ot_surgeries", str(surgery_id), {"surgery_id": surgery_id})
    return jsonify({"status": "ok"})


@app.get("/api/ot/summary")
@require_permissions("ot.read")
def ot_summary():
    return jsonify(get_ot_summary())


# ==================== Accounts ====================


@app.get("/api/accounts/summary")
@require_permissions("accounts.read")
def accounts_summary():
    return jsonify(get_accounts_summary())


@app.get("/api/accounts/ledger")
@require_permissions("accounts.read")
def accounts_ledger_list():
    entry_type = request.args.get("entry_type")
    return jsonify({"entries": rows_to_dicts(list_account_ledger_entries(entry_type=entry_type))})


@app.post("/api/accounts/ledger")
@require_permissions("accounts.write")
def accounts_ledger_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["entry_date", "entry_type", "category", "amount"])
    if validation_error:
        return validation_error
    entry_id = create_account_ledger_entry(payload)
    log_audit_event("create", "accounts_ledger", str(entry_id), {"category": payload.get("category")})
    return jsonify({"entry_id": entry_id})


@app.put("/api/accounts/ledger/<int:entry_id>")
@require_permissions("accounts.write")
def accounts_ledger_update(entry_id):
    payload = request.get_json(force=True)
    updated = update_account_ledger_entry(entry_id, payload)
    if not updated:
        return jsonify({"error": "Ledger entry not found"}), 404
    log_audit_event("update", "accounts_ledger", str(entry_id), {"entry_id": entry_id})
    return jsonify({"status": "ok"})


@app.delete("/api/accounts/ledger/<int:entry_id>")
@require_permissions("accounts.write")
def accounts_ledger_delete(entry_id):
    deleted = delete_account_ledger_entry(entry_id)
    if not deleted:
        return jsonify({"error": "Ledger entry not found"}), 404
    log_audit_event("delete", "accounts_ledger", str(entry_id), {"entry_id": entry_id})
    return jsonify({"status": "ok"})


@app.get("/api/accounts/vendors")
@require_permissions("accounts.read")
def accounts_vendor_payments_list():
    vendor_name = request.args.get("vendor_name")
    return jsonify({"payments": rows_to_dicts(list_vendor_payments(vendor_name=vendor_name))})


@app.post("/api/accounts/vendors")
@require_permissions("accounts.write")
def accounts_vendor_payments_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["vendor_name", "amount", "payment_date"])
    if validation_error:
        return validation_error
    payment_id = create_vendor_payment(payload)
    log_audit_event("create", "vendor_payments", str(payment_id), {"vendor_name": payload.get("vendor_name")})
    return jsonify({"payment_id": payment_id})


@app.put("/api/accounts/vendors/<int:payment_id>")
@require_permissions("accounts.write")
def accounts_vendor_payments_update(payment_id):
    payload = request.get_json(force=True)
    updated = update_vendor_payment(payment_id, payload)
    if not updated:
        return jsonify({"error": "Vendor payment not found"}), 404
    log_audit_event("update", "vendor_payments", str(payment_id), {"payment_id": payment_id})
    return jsonify({"status": "ok"})


@app.delete("/api/accounts/vendors/<int:payment_id>")
@require_permissions("accounts.write")
def accounts_vendor_payments_delete(payment_id):
    deleted = delete_vendor_payment(payment_id)
    if not deleted:
        return jsonify({"error": "Vendor payment not found"}), 404
    log_audit_event("delete", "vendor_payments", str(payment_id), {"payment_id": payment_id})
    return jsonify({"status": "ok"})


@app.get("/api/accounts/doctors")
@require_permissions("accounts.read")
def accounts_doctor_payouts_list():
    doctor_name = request.args.get("doctor_name")
    return jsonify({"payouts": rows_to_dicts(list_doctor_payouts(doctor_name=doctor_name))})


@app.post("/api/accounts/doctors")
@require_permissions("accounts.write")
def accounts_doctor_payouts_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["doctor_name", "payout_month", "amount"])
    if validation_error:
        return validation_error
    payout_id = create_doctor_payout(payload)
    log_audit_event("create", "doctor_payouts", str(payout_id), {"doctor_name": payload.get("doctor_name")})
    return jsonify({"payout_id": payout_id})


@app.put("/api/accounts/doctors/<int:payout_id>")
@require_permissions("accounts.write")
def accounts_doctor_payouts_update(payout_id):
    payload = request.get_json(force=True)
    updated = update_doctor_payout(payout_id, payload)
    if not updated:
        return jsonify({"error": "Doctor payout not found"}), 404
    log_audit_event("update", "doctor_payouts", str(payout_id), {"payout_id": payout_id})
    return jsonify({"status": "ok"})


@app.delete("/api/accounts/doctors/<int:payout_id>")
@require_permissions("accounts.write")
def accounts_doctor_payouts_delete(payout_id):
    deleted = delete_doctor_payout(payout_id)
    if not deleted:
        return jsonify({"error": "Doctor payout not found"}), 404
    log_audit_event("delete", "doctor_payouts", str(payout_id), {"payout_id": payout_id})
    return jsonify({"status": "ok"})


# ==================== HRMS ====================


@app.get("/api/hr/departments")
@require_permissions("hr.read")
def hr_departments_list():
    return jsonify({"departments": rows_to_dicts(list_departments(hospital_id=current_hospital_id()))})


@app.post("/api/hr/departments")
@require_permissions("hr.write")
def hr_departments_create():
    payload = request.get_json(force=True)
    try:
        department_name = normalize_department_name(payload.get("department_name"))
    except BadRequest as error:
        return jsonify({"error": str(error)}), 400

    existing = next(
        (
            row
            for row in rows_to_dicts(list_departments(hospital_id=current_hospital_id()))
            if (row.get("department_name") or "").strip().lower() == department_name.lower()
        ),
        None,
    )
    if existing:
        return jsonify({"department_id": existing.get("id"), "department_name": existing.get("department_name"), "already_exists": True})

    department_id = create_department(
        {
            "department_name": department_name,
            "mapped_head_employee_id": payload.get("mapped_head_employee_id"),
        },
        hospital_id=current_hospital_id(),
    )
    log_audit_event(
        "create",
        "departments",
        str(department_id),
        {"department_name": department_name},
    )
    return jsonify({"department_id": department_id, "department_name": department_name})


@app.put("/api/hr/departments/<int:department_id>")
@require_permissions("hr.write")
def hr_departments_update(department_id):
    payload = request.get_json(force=True)
    if "department_name" in payload:
        try:
            payload["department_name"] = normalize_department_name(payload.get("department_name"))
        except BadRequest as error:
            return jsonify({"error": str(error)}), 400

        duplicate = next(
            (
                row
                for row in rows_to_dicts(list_departments(hospital_id=current_hospital_id()))
                if row.get("id") != department_id
                and (row.get("department_name") or "").strip().lower() == payload["department_name"].lower()
            ),
            None,
        )
        if duplicate:
            return jsonify({"error": "Department already exists"}), 409

    updated = update_department(department_id, payload, hospital_id=current_hospital_id())
    if not updated:
        return jsonify({"error": "Department not found"}), 404
    log_audit_event(
        "update", "departments", str(department_id), {"department_id": department_id}
    )
    return jsonify({"status": "ok"})


@app.delete("/api/hr/departments/<int:department_id>")
@require_permissions("hr.write")
def hr_departments_delete(department_id):
    deleted = delete_department(department_id, hospital_id=current_hospital_id())
    if not deleted:
        return jsonify({"error": "Department not found"}), 404
    log_audit_event(
        "delete", "departments", str(department_id), {"department_id": department_id}
    )
    return jsonify({"status": "ok"})


@app.get("/api/hr/attendance")
@require_permissions("hr.read")
def hr_attendance_list():
    employee_id = request.args.get("employee_id")
    return jsonify(
        {"attendance": rows_to_dicts(list_attendance(employee_id=employee_id))}
    )


@app.post("/api/hr/attendance")
@require_permissions("hr.write")
def hr_attendance_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload, ["employee_id", "attendance_date", "status"]
    )
    if validation_error:
        return validation_error
    attendance_id = create_attendance(payload)
    log_audit_event(
        "create",
        "attendance",
        str(attendance_id),
        {"employee_id": payload.get("employee_id")},
    )
    return jsonify({"attendance_id": attendance_id})


@app.put("/api/hr/attendance/<int:attendance_id>")
@require_permissions("hr.write")
def hr_attendance_update(attendance_id):
    payload = request.get_json(force=True)
    updated = update_attendance_record(attendance_id, payload)
    if not updated:
        return jsonify({"error": "Attendance record not found"}), 404
    log_audit_event(
        "update", "attendance", str(attendance_id), {"attendance_id": attendance_id}
    )
    return jsonify({"status": "ok"})


@app.delete("/api/hr/attendance/<int:attendance_id>")
@require_permissions("hr.write")
def hr_attendance_delete(attendance_id):
    deleted = delete_attendance_record(attendance_id)
    if not deleted:
        return jsonify({"error": "Attendance record not found"}), 404
    log_audit_event(
        "delete", "attendance", str(attendance_id), {"attendance_id": attendance_id}
    )
    return jsonify({"status": "ok"})


@app.get("/api/hr/payroll")
@require_permissions("hr.read")
def hr_payroll_list():
    employee_id = request.args.get("employee_id")
    return jsonify({"payroll": rows_to_dicts(list_payroll(employee_id=employee_id))})


@app.post("/api/hr/payroll")
@require_permissions("hr.write")
def hr_payroll_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload, ["employee_id", "payroll_month", "basic_salary"]
    )
    if validation_error:
        return validation_error
    payroll_id = create_payroll_record(payload)
    log_audit_event(
        "create",
        "payroll",
        str(payroll_id),
        {"employee_id": payload.get("employee_id")},
    )
    return jsonify({"payroll_id": payroll_id})


@app.put("/api/hr/payroll/<int:payroll_id>")
@require_permissions("hr.write")
def hr_payroll_update(payroll_id):
    payload = request.get_json(force=True)
    updated = update_payroll_record(payroll_id, payload)
    if not updated:
        return jsonify({"error": "Payroll record not found"}), 404
    log_audit_event("update", "payroll", str(payroll_id), {"payroll_id": payroll_id})
    return jsonify({"status": "ok"})


@app.delete("/api/hr/payroll/<int:payroll_id>")
@require_permissions("hr.write")
def hr_payroll_delete(payroll_id):
    deleted = delete_payroll_record(payroll_id)
    if not deleted:
        return jsonify({"error": "Payroll record not found"}), 404
    log_audit_event("delete", "payroll", str(payroll_id), {"payroll_id": payroll_id})
    return jsonify({"status": "ok"})


@app.get("/api/hr/leaves")
@require_permissions("hr.read")
def hr_leaves_list():
    employee_id = request.args.get("employee_id")
    return jsonify(
        {"leaves": rows_to_dicts(list_leave_requests(employee_id=employee_id))}
    )


@app.post("/api/hr/leaves")
@require_permissions("hr.write")
def hr_leaves_create():
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(
        payload, ["employee_id", "leave_type", "start_date", "end_date"]
    )
    if validation_error:
        return validation_error
    leave_id = create_leave_request(payload)
    log_audit_event(
        "create",
        "leave_requests",
        str(leave_id),
        {"employee_id": payload.get("employee_id")},
    )
    return jsonify({"leave_id": leave_id})


@app.post("/api/hr/leaves/<int:leave_id>/status")
@require_permissions("hr.write")
def hr_leave_status_update(leave_id):
    payload = request.get_json(force=True)
    validation_error = validate_required_fields(payload, ["status"])
    if validation_error:
        return validation_error
    updated = update_leave_status(
        leave_id, payload.get("status", "pending"), g.current_user.get("username")
    )
    if not updated:
        return jsonify({"error": "Leave request not found"}), 404
    log_audit_event(
        "update", "leave_requests", str(leave_id), {"status": payload.get("status")}
    )
    return jsonify({"status": "ok"})


@app.delete("/api/hr/leaves/<int:leave_id>")
@require_permissions("hr.write")
def hr_leaves_delete(leave_id):
    deleted = delete_leave_request(leave_id)
    if not deleted:
        return jsonify({"error": "Leave request not found"}), 404
    log_audit_event("delete", "leave_requests", str(leave_id), {"leave_id": leave_id})
    return jsonify({"status": "ok"})


# ==================== Audit ====================


@app.get("/api/audit/logs")
@require_permissions("audit.read")
def audit_logs_list():
    module_name = request.args.get("module")
    limit = request.args.get("limit", default=100, type=int)
    return jsonify(
        {"logs": rows_to_dicts(get_audit_logs(module_name=module_name, limit=limit))}
    )


@app.get("/")
def serve_frontend_index():
    index_path = os.path.join(FRONTEND_DIST, "index.html")
    if os.path.exists(index_path):
        return send_from_directory(FRONTEND_DIST, "index.html")
    return jsonify({"error": "Frontend build not found. Run npm run build:frontend."}), 503


@app.get("/<path:frontend_path>")
def serve_frontend_asset_or_route(frontend_path):
    if frontend_path.startswith("api/"):
        return jsonify({"error": "API route not found"}), 404
    requested_path = os.path.join(FRONTEND_DIST, frontend_path)
    if os.path.exists(requested_path) and os.path.isfile(requested_path):
        return send_from_directory(FRONTEND_DIST, frontend_path)
    index_path = os.path.join(FRONTEND_DIST, "index.html")
    if os.path.exists(index_path):
        return send_from_directory(FRONTEND_DIST, "index.html")
    return jsonify({"error": "Frontend build not found. Run npm run build:frontend."}), 503


if __name__ == "__main__":
    host = os.getenv("FLASK_HOST") or os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT") or os.getenv("PORT", "5001"))
    print(f"[+] Flask running on {host}:{port}")
    app.run(host=host, port=port, debug=False, use_reloader=False)
